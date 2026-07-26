# -*- coding: utf-8 -*-
"""
表格截图识别 → Excel（独立于发票识别）

约定：每张图片生成一个同名 .xlsx（仅扩展名不同）。
"""

from __future__ import annotations

import os
import re
import sys
from typing import Callable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 复用现有 OCR 引擎（不重复加载模型）
from invoice_extractor import (
    IMAGE_EXTENSIONS,
    _get_ocr,
    is_supported_invoice_file,
    reset_ocr_if_failed,
    set_ocr_mode,
    normalize_ocr_mode,
    OCR_MODE_BALANCED,
)

# 期权/成交明细常见表头（续页无表头时用作默认列名）
DEFAULT_TRADE_HEADERS = [
    '发生日期', '成交时间', '合约代码', '合约名称', '证券代码',
    '持仓类别', '买卖', '开平', '备兑标志', '业务状态',
    '成交价格', '成交金额', '成交笔数', '发生数量', '后证券额',
]

HEADER_HINTS = ('发生日期', '成交时间', '合约代码', '合约名称', '证券代码')

BoxItem = Tuple[str, float, float, float, float]  # text, x1, y1, x2, y2


def is_table_image_file(path: str) -> bool:
    ext = os.path.splitext(path)[1].lower()
    return ext in IMAGE_EXTENSIONS


def _box_from_poly(poly) -> Optional[Tuple[float, float, float, float]]:
    if poly is None:
        return None
    try:
        xs, ys = [], []
        for pt in poly:
            if isinstance(pt, (list, tuple)) and len(pt) >= 2:
                xs.append(float(pt[0]))
                ys.append(float(pt[1]))
            elif hasattr(pt, '__len__') and len(pt) >= 2:
                xs.append(float(pt[0]))
                ys.append(float(pt[1]))
        if not xs or not ys:
            return None
        return min(xs), min(ys), max(xs), max(ys)
    except Exception:
        return None


def _extract_items_from_ocr_result(result) -> List[BoxItem]:
    """从 PaddleOCR 2.x/3.x 结果中提取 (text, x1,y1,x2,y2)。"""
    items: List[BoxItem] = []
    if not result:
        return items

    pages = result if isinstance(result, (list, tuple)) else [result]
    for page in pages:
        if page is None:
            continue

        # PaddleOCR 3.x dict / result object
        texts = None
        polys = None
        if isinstance(page, dict):
            texts = page.get('rec_texts') or page.get('texts')
            polys = page.get('rec_polys') or page.get('dt_polys') or page.get('rec_boxes')
        elif hasattr(page, 'get'):
            try:
                texts = page.get('rec_texts') or page.get('texts')
                polys = page.get('rec_polys') or page.get('dt_polys') or page.get('rec_boxes')
            except Exception:
                pass
        if (not texts) and hasattr(page, 'json'):
            try:
                payload = page.json() if callable(page.json) else page.json
                if isinstance(payload, dict):
                    res = payload.get('res', payload)
                    if isinstance(res, dict):
                        texts = res.get('rec_texts') or texts
                        polys = res.get('rec_polys') or res.get('dt_polys') or polys
            except Exception:
                pass

        if texts and polys is not None:
            try:
                poly_list = list(polys)
            except Exception:
                poly_list = polys
            for i, text in enumerate(texts):
                t = str(text or '').strip()
                if not t:
                    continue
                poly = poly_list[i] if i < len(poly_list) else None
                # rec_boxes 可能是 [x1,y1,x2,y2]
                box = None
                if poly is not None and hasattr(poly, '__len__') and len(poly) == 4 and not isinstance(poly[0], (list, tuple)):
                    try:
                        x1, y1, x2, y2 = map(float, poly)
                        box = (x1, y1, x2, y2)
                    except Exception:
                        box = None
                if box is None:
                    box = _box_from_poly(poly)
                if box is None:
                    continue
                items.append((t, box[0], box[1], box[2], box[3]))
            if items:
                continue

        # PaddleOCR 2.x: [[[box], (text, score)], ...]
        if isinstance(page, (list, tuple)):
            for line in page:
                if not isinstance(line, (list, tuple)) or len(line) < 2:
                    continue
                box_raw, text_part = line[0], line[1]
                if isinstance(text_part, (list, tuple)) and text_part:
                    t = str(text_part[0]).strip()
                elif isinstance(text_part, str):
                    t = text_part.strip()
                else:
                    continue
                if not t:
                    continue
                box = _box_from_poly(box_raw)
                if box:
                    items.append((t, box[0], box[1], box[2], box[3]))

    return items


def _ocr_items(image_path: str) -> List[BoxItem]:
    ocr = _get_ocr()
    if ocr is None:
        raise RuntimeError('OCR 引擎不可用')

    result = None
    if hasattr(ocr, 'predict'):
        try:
            result = ocr.predict(image_path)
        except Exception as e:
            sys.stderr.write(f'  [表格OCR] predict 失败: {e}\n')
            sys.stderr.flush()
    if result is None and hasattr(ocr, 'ocr'):
        try:
            result = ocr.ocr(image_path, cls=False)
        except TypeError:
            result = ocr.ocr(image_path)
        except Exception as e:
            sys.stderr.write(f'  [表格OCR] ocr() 失败: {e}\n')
            sys.stderr.flush()

    items = _extract_items_from_ocr_result(result)
    if not items:
        # 退化：仅有文字、无坐标时按行切分（单列）
        from invoice_extractor import _run_ocr_on_image
        text = _run_ocr_on_image(ocr, image_path)
        rows = []
        for i, line in enumerate((text or '').splitlines()):
            line = line.strip()
            if line:
                rows.append((line, 0.0, float(i), 100.0, float(i) + 0.8))
        return rows
    return items


def _cluster_rows(items: List[BoxItem], y_tol_ratio: float = 0.45) -> List[List[BoxItem]]:
    if not items:
        return []
    heights = [max(4.0, y2 - y1) for _, _, y1, _, y2 in items]
    median_h = sorted(heights)[len(heights) // 2]
    y_tol = max(6.0, median_h * y_tol_ratio)

    ordered = sorted(items, key=lambda it: ((it[2] + it[4]) / 2.0, it[1]))
    rows: List[List[BoxItem]] = []
    row_ys: List[float] = []

    for it in ordered:
        cy = (it[2] + it[4]) / 2.0
        if not rows:
            rows.append([it])
            row_ys.append(cy)
            continue
        if abs(cy - row_ys[-1]) <= y_tol:
            rows[-1].append(it)
            row_ys[-1] = sum((x[2] + x[4]) / 2.0 for x in rows[-1]) / len(rows[-1])
        else:
            rows.append([it])
            row_ys.append(cy)

    merged_rows = []
    for row in rows:
        row = sorted(row, key=lambda it: it[1])
        merged_rows.append(_merge_horizontal(row))
    return merged_rows


def _merge_horizontal(row: List[BoxItem], gap_ratio: float = 0.18) -> List[BoxItem]:
    """只合并明显属于同一单元格的碎片；默认偏保守，避免跨列粘连。"""
    if not row:
        return []
    widths = [max(4.0, x2 - x1) for _, x1, _, x2, _ in row]
    median_w = sorted(widths)[len(widths) // 2]
    gap = max(4.0, median_w * gap_ratio)

    out: List[BoxItem] = []
    cur_t, x1, y1, x2, y2 = row[0]
    for t, a, b, c, d in row[1:]:
        close = (a - x2) <= gap
        if close and not _looks_like_separate_cells(cur_t, t):
            cur_t = cur_t + t
            x2 = max(x2, c)
            y1 = min(y1, b)
            y2 = max(y2, d)
        else:
            out.append((cur_t, x1, y1, x2, y2))
            cur_t, x1, y1, x2, y2 = t, a, b, c, d
    out.append((cur_t, x1, y1, x2, y2))
    return out


# 买卖/开平 OCR 误读表（短标签 + 粘连长串内替换）
_BS_OCR_FIX = {
    '续出': '卖出', '读出': '卖出', '游出': '卖出', '麦出': '卖出',
    '荚出': '卖出', '卖山': '卖出', '卖士': '卖出', '卖土': '卖出',
    '买人': '买入', '买八': '买入',
}
_OC_OCR_FIX = {
    '升仓': '开仓', '苹仓': '平仓',
}


def _looks_like_separate_cells(left: str, right: str) -> bool:
    """避免把相邻两列误合并。"""
    left = (left or '').strip()
    right = (right or '').strip()
    date_re = re.compile(r'^\d{8}$')
    time_re = re.compile(r'^\d{1,2}:\d{2}(:\d{2})?$')
    num_re = re.compile(r'^-?\d+(\.\d+)?$')
    code6 = re.compile(r'^\d{6}$')
    code8 = re.compile(r'^\d{8}$')
    cats = {
        '买入', '卖出', '开仓', '平仓', '义务方', '权利方', '非备兑', '备兑', '成交',
        *_BS_OCR_FIX.keys(), *_OC_OCR_FIX.keys(),
    }

    if left in cats or right in cats:
        return True
    if date_re.match(left) or date_re.match(right):
        return True
    if time_re.match(left) or time_re.match(right):
        return True
    # 证券代码 6 位、合约代码 8 位绝不与邻格合并
    if code6.match(left) or code6.match(right) or code8.match(left) or code8.match(right):
        return True
    if num_re.match(left) and num_re.match(right):
        return True
    # 合约名称末尾常是数字（如 3000），右侧是证券代码
    if re.search(r'\d$', left) and code6.match(right):
        return True
    if re.search(r'[\u4e00-\u9fffA-Za-z]$', left) and num_re.match(right):
        return True
    if num_re.match(left) and re.match(r'^[\u4e00-\u9fffA-Za-z]', right):
        return True
    # 名称片段 + 权利方/义务方
    if right in ('权利方', '义务方') or left in ('权利方', '义务方'):
        return True
    return False


def _row_texts(row: List[BoxItem]) -> List[str]:
    return [t for t, *_ in row]


def _is_header_row(texts: List[str]) -> bool:
    joined = ''.join(texts)
    hits = sum(1 for h in HEADER_HINTS if h in joined)
    return hits >= 2


def _looks_like_trade_row(texts: List[str]) -> bool:
    if not texts:
        return False
    joined = ''.join(texts)
    if re.search(r'\d{8}', texts[0] if texts else ''):
        return True
    if 'ETF' in joined or '沽' in joined or '购' in joined or '洁' in joined or '活' in joined:
        return True
    if any(x in joined for x in ('义务方', '权利方', '开仓', '平仓', '非备兑')):
        return True
    return False


def _fix_ocr_chars(s: str) -> str:
    """常见 OCR 误读修正。"""
    if not s:
        return s
    s = str(s).strip()
    if s in _BS_OCR_FIX:
        return _BS_OCR_FIX[s]
    if s in _OC_OCR_FIX:
        return _OC_OCR_FIX[s]
    # 模糊：两字且含「出」、不像其他标签 → 卖出
    if len(s) == 2 and '出' in s and s not in ('非备兑', '成交') and '买' not in s:
        if '卖' in s or s[0] in ('续', '读', '游', '麦', '荚', '卖'):
            return '卖出'
    if len(s) == 2 and '买' in s and '仓' not in s:
        return '买入'
    # 粘连长串内替换（整行正则兜底前）
    for bad, good in _BS_OCR_FIX.items():
        if bad in s:
            s = s.replace(bad, good)
    for bad, good in _OC_OCR_FIX.items():
        if bad in s:
            s = s.replace(bad, good)
    # 沽 常被识别成 洁/活
    s = re.sub(r'(ETF)洁', r'\1沽', s)
    s = re.sub(r'(ETF)活', r'\1沽', s)
    s = s.replace('√', '1').replace('、', '1')
    return s


def _split_glued_name_code(text: str) -> Tuple[str, str, str]:
    """
    拆开粘连的：合约代码 + 合约名称 + 证券代码
    返回 (code8_or_empty, name, sec6_or_empty)
    """
    t = _fix_ocr_chars((text or '').strip())
    if not t:
        return '', '', ''

    m = re.match(r'^(\d{8})(.+?)(\d{6})$', t)
    if m:
        return m.group(1), m.group(2), m.group(3)

    m = re.match(r'^(\d{8})(.+)$', t)
    if m and re.search(r'[\u4e00-\u9fff]', m.group(2)):
        name = m.group(2)
        m2 = re.match(r'^(.+?)(\d{6})$', name)
        if m2 and ('ETF' in m2.group(1) or '沽' in m2.group(1) or '购' in m2.group(1)):
            return m.group(1), m2.group(1), m2.group(2)
        return m.group(1), name, ''

    m = re.match(r'^(.+?)(\d{6})$', t)
    if m and ('ETF' in m.group(1) or re.search(r'[沽购洁活]', m.group(1))):
        return '', m.group(1), m.group(2)

    return '', t, ''


def _estimate_col_centers(box_rows: List[List[BoxItem]], expect_cols: int = 15) -> Optional[List[float]]:
    """用接近完整列数的行估计各列中心 x。"""
    candidates = [r for r in box_rows if abs(len(r) - expect_cols) <= 2 and len(r) >= expect_cols - 3]
    if not candidates:
        candidates = [r for r in box_rows if len(r) >= max(10, expect_cols - 5)]
    if not candidates:
        return None

    candidates = sorted(candidates, key=lambda r: abs(len(r) - expect_cols))[:8]
    exact = [r for r in candidates if len(r) == expect_cols]
    if exact:
        centers = []
        for i in range(expect_cols):
            xs = [((row[i][1] + row[i][3]) / 2.0) for row in exact]
            centers.append(sum(xs) / len(xs))
        return centers

    best = max(candidates, key=len)
    xs = [((it[1] + it[3]) / 2.0) for it in best]
    if len(xs) < 3:
        return None
    if len(xs) == expect_cols:
        return xs
    out = []
    for i in range(expect_cols):
        pos = i * (len(xs) - 1) / (expect_cols - 1)
        lo = int(pos)
        hi = min(lo + 1, len(xs) - 1)
        frac = pos - lo
        out.append(xs[lo] * (1 - frac) + xs[hi] * frac)
    return out


def _assign_row_to_columns(row: List[BoxItem], centers: List[float]) -> List[str]:
    cols = ['' for _ in centers]
    for t, x1, y1, x2, y2 in row:
        cx = (x1 + x2) / 2.0
        j = min(range(len(centers)), key=lambda i: abs(centers[i] - cx))
        if cols[j]:
            cols[j] = cols[j] + t
        else:
            cols[j] = t
    return cols


def _is_well_formed_trade_row(cells: List[str]) -> bool:
    if len(cells) != 15:
        return False
    if not re.match(r'^\d{8}$', cells[0] or ''):
        return False
    if not re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', cells[1] or ''):
        return False
    if not re.match(r'^\d{8}$', cells[2] or ''):
        return False
    if not (cells[3] or '').strip():
        return False
    # 名称不应尾粘 6 位证券代码
    if re.search(r'\d{6}$', cells[3] or '') and ('ETF' in (cells[3] or '') or re.search(r'[沽购洁活]', cells[3] or '')):
        if not re.match(r'^\d{6}$', cells[4] or ''):
            return False
    if not re.match(r'^\d{6}$', cells[4] or ''):
        return False
    if cells[5] not in ('义务方', '权利方'):
        return False
    if cells[6] not in ('买入', '卖出'):
        return False
    if cells[7] not in ('开仓', '平仓'):
        return False
    if cells[8] not in ('非备兑', '备兑'):
        return False
    if cells[9] != '成交':
        return False
    # 价格/金额形态
    if not re.match(r'^\d+(\.\d+)?$', cells[10] or ''):
        return False
    if not re.match(r'^\d+(\.\d+)?$', cells[11] or ''):
        return False
    # 笔数允许 √ 等，后面再修
    if not re.match(r'^(\d+|√|、)$', str(cells[12] or '')):
        return False
    if not re.match(r'^\d+(\.\d+)?$', cells[13] or ''):
        return False
    if not re.match(r'^\d+(\.\d+)?$', cells[14] or ''):
        return False
    return True


def _parse_trade_line(joined: str) -> Optional[List[str]]:
    """整行正则兜底解析（粘连严重时）。"""
    s = _fix_ocr_chars(re.sub(r'\s+', '', joined or ''))
    if not s:
        return None
    pat = (
        r'^(?P<date>\d{8})'
        r'(?P<time>\d{1,2}:\d{2}:\d{2})'
        r'(?P<code>\d{8})'
        r'(?P<name>.+?)'
        r'(?P<sec>\d{6})'
        r'(?P<pos>义务方|权利方)'
        r'(?P<bs>买入|卖出|续出|读出|游出|麦出|荚出|卖山|卖士|卖土|买人|买八)'
        r'(?P<oc>开仓|平仓|升仓|苹仓)'
        r'(?P<cov>非备兑|备兑)'
        r'(?P<st>成交)'
        r'(?P<price>0\.\d{3,4})'
        r'(?P<amt>\d{1,8}\.\d{2})'
        r'(?P<cnt>\d{1,2})'
        r'(?P<qty>\d{1,8}\.\d{2})'
        r'(?P<bal>\d{1,8}\.\d{2})$'
    )
    m = re.match(pat, s)
    if not m:
        return None
    bs = _BS_OCR_FIX.get(m.group('bs'), m.group('bs'))
    oc = _OC_OCR_FIX.get(m.group('oc'), m.group('oc'))
    return [
        m.group('date'), m.group('time'), m.group('code'), m.group('name'), m.group('sec'),
        m.group('pos'), bs, oc, m.group('cov'), m.group('st'),
        m.group('price'), m.group('amt'), m.group('cnt'), m.group('qty'), m.group('bal'),
    ]


def _polish_trade_row(cells: List[str]) -> List[str]:
    out = [_fix_ocr_chars(str(c or '').strip()) for c in cells]
    if len(out) != 15:
        return out
    # 名称尾粘证券代码
    code8, name, sec6 = _split_glued_name_code(out[3])
    if sec6 and not out[4]:
        out[3] = name
        out[4] = sec6
    elif sec6 and out[4] == sec6:
        out[3] = name
    if code8 and not out[2]:
        out[2] = code8
        out[3] = name or out[3]
    # 笔数 OCR 噪声
    if out[12] in ('√', '、', 'l', 'I', '|'):
        out[12] = '1'
    # 价格被粘金额：0.0048480.00
    m = re.match(r'^(0\.\d{3,4})(\d+\.\d{2})$', out[10] or '')
    if m and (not out[11] or out[11] == m.group(2)):
        out[10], out[11] = m.group(1), m.group(2)
    # 截断行常见缺省：备兑/状态几乎总是这两个值
    if out[5] in ('义务方', '权利方') and out[10]:
        if not out[8]:
            out[8] = '非备兑'
        if not out[9]:
            out[9] = '成交'
    return out


def _repair_trade_row(cells: List[str]) -> List[str]:
    """把识别结果校正为标准 15 列。优先保留已对齐的 15 列，避免二次拆坏。"""
    raw = [_fix_ocr_chars(str(c or '').strip()) for c in cells]

    # 列对齐成功的行：只做轻量抛光
    if len(raw) == 15 and _is_well_formed_trade_row(raw):
        return _polish_trade_row(raw)

    # 若末尾 5 个已是价格/金额/笔数/数量/余额，只修前面字段
    compact = [c for c in raw if c != '']
    if len(compact) >= 6:
        tail = compact[-5:]
        if (
            re.match(r'^0\.\d{3,4}$', tail[0] or '')
            and re.match(r'^\d+\.\d{2}$', tail[1] or '')
            and re.match(r'^(\d{1,3}|√|、)$', tail[2] or '')
            and re.match(r'^\d+\.\d{2}$', tail[3] or '')
            and re.match(r'^\d+\.\d{2}$', tail[4] or '')
        ):
            head = compact[:-5]
            expanded: List[str] = []
            for c in head:
                code8, name, sec6 = _split_glued_name_code(c)
                if code8 or sec6:
                    if code8:
                        expanded.append(code8)
                    if name:
                        expanded.append(name)
                    if sec6:
                        expanded.append(sec6)
                else:
                    expanded.append(c)
            # 拼成 15 列：expanded 应约 10 个字段
            cnt = '1' if tail[2] in ('√', '、') else tail[2]
            # 用分类字段填充
            out = [''] * 15
            i = 0
            n = len(expanded)

            def take():
                nonlocal i
                if i < n:
                    v = expanded[i]
                    i += 1
                    return v
                return ''

            if i < n and re.match(r'^\d{8}$', expanded[i]):
                out[0] = take()
            if i < n and re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', expanded[i]):
                out[1] = take()
            if i < n and re.match(r'^\d{8}$', expanded[i]):
                out[2] = take()
            name_parts = []
            while i < n and not re.match(r'^\d{6}$', expanded[i]) and expanded[i] not in ('义务方', '权利方'):
                if expanded[i] in ('买入', '卖出', '开仓', '平仓', '非备兑', '成交'):
                    break
                name_parts.append(take())
                if len(name_parts) >= 3:
                    break
            out[3] = ''.join(name_parts)
            if i < n and re.match(r'^\d{6}$', expanded[i]):
                out[4] = take()
            for idx, options in [
                (5, ('义务方', '权利方')),
                (6, ('买入', '卖出')),
                (7, ('开仓', '平仓')),
                (8, ('非备兑', '备兑')),
                (9, ('成交',)),
            ]:
                if i < n and expanded[i] in options:
                    out[idx] = take()
            out[10], out[11], out[12], out[13], out[14] = tail[0], tail[1], cnt, tail[3], tail[4]
            if out[0] and out[2] and out[4]:
                return _polish_trade_row(out)

    cells = compact
    parsed = _parse_trade_line(''.join(cells))
    if parsed:
        return _polish_trade_row(parsed)

    expanded = []
    for c in cells:
        code8, name, sec6 = _split_glued_name_code(c)
        if code8 or sec6:
            if code8:
                expanded.append(code8)
            if name:
                expanded.append(name)
            if sec6:
                expanded.append(sec6)
        else:
            expanded.append(c)

    parsed = _parse_trade_line(''.join(expanded))
    if parsed:
        return _polish_trade_row(parsed)

    out = [''] * 15
    i = 0
    n = len(expanded)

    def take2():
        nonlocal i
        if i < n:
            v = expanded[i]
            i += 1
            return v
        return ''

    if i < n and re.match(r'^\d{8}$', expanded[i]):
        out[0] = take2()
    if i < n and re.match(r'^\d{1,2}:\d{2}(:\d{2})?$', expanded[i]):
        out[1] = take2()
    if i < n and re.match(r'^\d{8}$', expanded[i]):
        out[2] = take2()

    name_parts = []
    while i < n and not re.match(r'^\d{6}$', expanded[i]) and expanded[i] not in ('义务方', '权利方'):
        if re.match(r'^\d+\.\d+$', expanded[i]) or expanded[i] in ('买入', '卖出', '开仓', '平仓', '非备兑', '成交'):
            break
        name_parts.append(take2())
        if len(name_parts) >= 3:
            break
    out[3] = ''.join(name_parts)

    if i < n and re.match(r'^\d{6}$', expanded[i]):
        out[4] = take2()
    for idx, options in [
        (5, ('义务方', '权利方')),
        (6, ('买入', '卖出')),
        (7, ('开仓', '平仓')),
        (8, ('非备兑', '备兑')),
        (9, ('成交',)),
    ]:
        if i < n and expanded[i] in options:
            out[idx] = take2()

    nums = []
    while i < n:
        v = take2()
        v = '1' if v in ('√', '、') else v
        if re.match(r'^-?\d+(\.\d+)?$', v):
            nums.append(v)
    if len(nums) >= 5:
        out[10:15] = nums[:5]
    elif len(nums) == 4:
        out[10], out[11], out[13], out[14] = nums
        out[12] = '1'
    else:
        for j, v in enumerate(nums[:5]):
            out[10 + j] = v

    if not out[0] or not out[2]:
        parsed = _parse_trade_line(''.join(expanded))
        if parsed:
            return _polish_trade_row(parsed)
    return _polish_trade_row(out)


def _normalize_grid(rows: List[List[str]]) -> Tuple[List[str], List[List[str]]]:
    header = list(DEFAULT_TRADE_HEADERS)
    data = []
    for r in rows:
        if not r or not any(str(c).strip() for c in r):
            continue
        if _is_header_row(r):
            continue
        if not _looks_like_trade_row(r):
            # 可能是噪声行
            joined = ''.join(r)
            if not re.search(r'\d{8}', joined):
                continue
        fixed = _repair_trade_row(r)
        # 过滤仍无日期的无效行
        if not re.match(r'^\d{8}$', fixed[0] or ''):
            continue
        data.append(fixed)
    return header, data


def image_to_table(image_path: str) -> Tuple[List[str], List[List[str]]]:
    """识别单张表格图片，返回 (headers, rows)。"""
    items = _ocr_items(image_path)
    box_rows = _cluster_rows(items)

    centers = _estimate_col_centers(box_rows, expect_cols=15)
    text_rows: List[List[str]] = []
    for row in box_rows:
        raw_texts = _row_texts(row)
        if centers:
            assigned = [_fix_ocr_chars(c) for c in _assign_row_to_columns(row, centers)]
            if _is_well_formed_trade_row(assigned):
                text_rows.append(assigned)
            elif len(raw_texts) >= 12:
                # 列对齐若把合约代码吸进名称，回退原始切分再交给 repair
                text_rows.append([_fix_ocr_chars(c) for c in raw_texts])
            else:
                text_rows.append(assigned)
        else:
            text_rows.append([_fix_ocr_chars(c) for c in raw_texts])

    return _normalize_grid(text_rows)


def write_table_excel(headers: List[str], rows: List[List[str]], output_path: str) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or '.', exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = '表格识别'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F6EF7')
    thin = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD'),
    )
    align = Alignment(vertical='center', wrap_text=True)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align
        cell.border = thin

    for r_i, row in enumerate(rows, 2):
        for c_i, val in enumerate(row, 1):
            cell = ws.cell(r_i, c_i, val)
            cell.alignment = align
            cell.border = thin

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = 14

    wb.save(output_path)
    return output_path


def image_file_to_excel(image_path: str, output_path: str = None) -> dict:
    """
    一张图片 → 一个同名 Excel。
    返回 {ok, source, excel, rows, cols, error}
    """
    stem = os.path.splitext(os.path.basename(image_path))[0]
    if output_path is None:
        output_path = os.path.join(os.path.dirname(image_path), stem + '.xlsx')

    try:
        headers, rows = image_to_table(image_path)
        write_table_excel(headers, rows, output_path)
        return {
            'ok': True,
            'source': os.path.basename(image_path),
            'excel': os.path.basename(output_path),
            'excel_path': output_path,
            'rows': len(rows),
            'cols': len(headers),
            'error': '',
        }
    except Exception as e:
        return {
            'ok': False,
            'source': os.path.basename(image_path),
            'excel': '',
            'excel_path': '',
            'rows': 0,
            'cols': 0,
            'error': str(e),
        }


def process_table_images(
    input_dir: str,
    output_dir: str,
    ocr_mode: str = OCR_MODE_BALANCED,
    progress_cb: Optional[Callable[[str, int], None]] = None,
) -> dict:
    """
    扫描目录中的图片，每张输出同名 xlsx 到 output_dir（保留相对子目录可选扁平化）。
    """
    reset_ocr_if_failed()
    set_ocr_mode(normalize_ocr_mode(ocr_mode))

    images = []
    for root, dirnames, filenames in os.walk(input_dir):
        dirnames[:] = [d for d in dirnames if not d.startswith('.') and d != '.chunks']
        for name in sorted(filenames):
            path = os.path.join(root, name)
            if is_table_image_file(path) or is_supported_invoice_file(path):
                # 表格模式只处理图片；误传 PDF 跳过
                if is_table_image_file(path):
                    images.append(path)

    os.makedirs(output_dir, exist_ok=True)
    results = []
    total = len(images)
    if total == 0:
        return {'ok': False, 'error': '未找到图片文件', 'files': [], 'success': 0, 'total': 0}

    # 同名冲突时加目录前缀
    used_names = set()
    for i, img in enumerate(images):
        if progress_cb:
            progress_cb(f'识别表格 ({i + 1}/{total}): {os.path.basename(img)}', int(i / total * 100))

        stem = os.path.splitext(os.path.basename(img))[0]
        out_name = stem + '.xlsx'
        if out_name.lower() in used_names:
            # 不同子目录下同名时加父目录名
            parent = os.path.basename(os.path.dirname(img))
            out_name = f'{parent}_{stem}.xlsx'
        used_names.add(out_name.lower())
        out_path = os.path.join(output_dir, out_name)

        info = image_file_to_excel(img, out_path)
        results.append(info)

    success = sum(1 for r in results if r.get('ok'))
    if progress_cb:
        progress_cb(f'完成：成功 {success}/{total}', 100)
    return {
        'ok': success > 0,
        'error': '' if success > 0 else '全部识别失败',
        'files': results,
        'success': success,
        'total': total,
    }
