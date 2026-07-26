"""
电子发票（普通发票）PDF 自动识别与 Excel 导出工具  v2.0
--------------------------------------------------------
功能：
  1. 自动扫描文件夹下所有 PDF 发票文件
  2. 智能判断 PDF 类型：
     - 文本型 PDF → pdfplumber 快速解析（精准）
     - 图片型 PDF → PaddleOCR 识别（支持中文）
  3. 智能判断单张/多张发票：
     - 单页 PDF → 1 张发票
     - 多页 PDF → 每页识别为独立发票
  4. 提取字段：发票号码、开票日期、购买方名称、销售方名称、合计金额、价税合计
  5. 导出 Excel：发票明细 + 报销汇总 两个 Sheet

用法：
  python invoice_extractor.py
  将脚本与 PDF 发票放在同一目录，或在 CONFIG 中配置路径。

依赖（自动检测，缺失时给出提示）：
  pip install pdfplumber openpyxl PyMuPDF paddlepaddle paddleocr
"""

import os
import re
import sys
import io
import tempfile
from pathlib import Path

# 抑制 PaddlePaddle / oneDNN 在 CPU 上 OCR 时刷屏的内部日志
# （如 ReduceMeanCheckIfOneDNNSupport，并非报错）
os.environ.setdefault('GLOG_minloglevel', '2')
os.environ.setdefault('FLAGS_min_log_level', '2')
os.environ.setdefault('MKLDNN_VERBOSE', '0')
os.environ.setdefault('PADDLEOCR_LOG_LEVEL', 'ERROR')

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pdfplumber
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 脚本所在目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ========================== 配置区 ==========================
CONFIG = {
    # PDF 发票所在的文件夹路径（None = 使用脚本所在目录）
    "input_dir": None,

    # Excel 输出路径（None = 输出到脚本目录下的 excel 子文件夹）
    "output_file": None,
}
# ============================================================

# 支持的发票文件扩展名
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.webp', '.tif', '.tiff'}
INVOICE_EXTENSIONS = {'.pdf'} | IMAGE_EXTENSIONS


def is_image_file(path: str) -> bool:
    """判断是否为发票图片文件。"""
    return os.path.splitext(path)[1].lower() in IMAGE_EXTENSIONS


def is_supported_invoice_file(path: str) -> bool:
    """判断是否为支持识别的发票文件（PDF 或图片）。"""
    return os.path.splitext(path)[1].lower() in INVOICE_EXTENSIONS

_ocr_instance = None
_ocr_failed = False
_ocr_instance_mode = None  # 当前实例对应的识别模式
_ocr_quality_mode = 'balanced'  # fast / balanced / accurate
_progress_callback = None

OCR_MODE_FAST = 'fast'
OCR_MODE_BALANCED = 'balanced'
OCR_MODE_ACCURATE = 'accurate'
OCR_MODES = (OCR_MODE_FAST, OCR_MODE_BALANCED, OCR_MODE_ACCURATE)

# 模式参数：模型档位、渲染 DPI、最大边、是否字段不全时重试
_OCR_MODE_PROFILES = {
    OCR_MODE_FAST: {
        'label': '速度',
        'use_server_model': False,
        'dpi_cpu': 140,
        'dpi_gpu': 180,
        'max_side_cpu': 1800,
        'max_side_gpu': 2400,
        'retry_incomplete': False,
    },
    OCR_MODE_BALANCED: {
        'label': '平衡',
        'use_server_model': False,
        'dpi_cpu': 160,
        'dpi_gpu': 220,
        'max_side_cpu': 2200,
        'max_side_gpu': 3200,
        'retry_incomplete': True,
    },
    OCR_MODE_ACCURATE: {
        'label': '精度',
        'use_server_model': True,
        'dpi_cpu': 180,
        'dpi_gpu': 260,
        'max_side_cpu': 2400,
        'max_side_gpu': 3900,
        'retry_incomplete': True,
        'retry_dpi_gpu': 320,
    },
}


def normalize_ocr_mode(mode) -> str:
    """规范化识别模式，非法值回退为平衡。"""
    m = (mode or '').strip().lower()
    aliases = {
        'speed': OCR_MODE_FAST,
        '快速': OCR_MODE_FAST,
        '速度': OCR_MODE_FAST,
        'balance': OCR_MODE_BALANCED,
        'balanced': OCR_MODE_BALANCED,
        '平衡': OCR_MODE_BALANCED,
        'accurate': OCR_MODE_ACCURATE,
        'accuracy': OCR_MODE_ACCURATE,
        'precision': OCR_MODE_ACCURATE,
        '精度': OCR_MODE_ACCURATE,
        '精准': OCR_MODE_ACCURATE,
    }
    if m in OCR_MODES:
        return m
    return aliases.get(m, OCR_MODE_BALANCED)


def get_ocr_mode() -> str:
    return _ocr_quality_mode


def set_ocr_mode(mode: str) -> str:
    """设置识别模式；若模型档位变化则重建 OCR 实例。"""
    global _ocr_quality_mode, _ocr_instance, _ocr_instance_mode, _ocr_failed
    new_mode = normalize_ocr_mode(mode)
    old_mode = _ocr_quality_mode
    _ocr_quality_mode = new_mode
    old_prof = _OCR_MODE_PROFILES.get(old_mode, _OCR_MODE_PROFILES[OCR_MODE_BALANCED])
    new_prof = _OCR_MODE_PROFILES[new_mode]
    # server/mobile 切换时必须换实例
    if (
        _ocr_instance is not None
        and bool(old_prof.get('use_server_model')) != bool(new_prof.get('use_server_model'))
    ):
        _ocr_instance = None
        _ocr_instance_mode = None
        _ocr_failed = False
        sys.stderr.write(f"  [模式] 切换 {old_mode} → {new_mode}，将重新加载 OCR 模型\n")
        sys.stderr.flush()
    else:
        sys.stderr.write(f"  [模式] 识别模式: {new_prof['label']} ({new_mode})\n")
        sys.stderr.flush()
    return new_mode


def _current_ocr_profile() -> dict:
    return _OCR_MODE_PROFILES.get(_ocr_quality_mode, _OCR_MODE_PROFILES[OCR_MODE_BALANCED])


def set_progress_callback(callback):
    """设置进度回调（供 Web 界面显示），签名: callback(message, page_progress=None)"""
    global _progress_callback
    _progress_callback = callback


def _emit_progress(message: str, page_progress=None):
    """输出进度到控制台（stderr）并通知回调。app.py 会重定向 stdout，必须用 stderr。"""
    sys.stderr.write(message + '\n')
    sys.stderr.flush()
    if _progress_callback:
        try:
            _progress_callback(message, page_progress)
        except TypeError:
            _progress_callback(message)


def reset_ocr_if_failed():
    """OCR 曾失败后，允许新批次重新尝试加载（如用户已清理缓存）。"""
    global _ocr_instance, _ocr_failed, _ocr_instance_mode
    if _ocr_failed:
        _ocr_instance = None
        _ocr_instance_mode = None
        _ocr_failed = False


def get_paddle_cache_dirs():
    """返回所有可能的 PaddleOCR / PaddleX 缓存目录。"""
    user_home = os.path.expanduser('~')
    dirs = [
        os.path.join(user_home, '.paddlex'),
        os.path.join(user_home, '.paddleocr'),
    ]
    if any(ord(c) > 127 for c in user_home):
        for drive in ('D:', 'E:', 'F:', 'C:'):
            if os.path.exists(drive + '\\'):
                dirs.append(os.path.join(drive + '\\', 'paddle_ocr_cache', '.paddlex'))
                break
    return dirs


def _safe_float(value, default=0.0):
    """安全转换为浮点数，避免 OCR 异常文本导致崩溃。"""
    if value is None or value == '':
        return default
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return default


def _buyer_name_base(name):
    """提取购买方名称（去掉括号及内容）。"""
    return re.sub(r'[（(].*[）)]', '', name or '').strip()


def _ensure_torch_before_paddle():
    """PaddleOCR 3.x → PaddleX → modelscope 会间接 import torch。
    若先加载 paddle 再加载 torch，Windows 上常见 shm.dll WinError 127。
    必须在任何 paddle 导入之前先加载 torch。"""
    try:
        import torch  # noqa: F401
    except Exception:
        # torch 缺失时仍尝试继续；真正失败会在 paddleocr 导入阶段暴露
        pass


def _nvidia_gpu_visible() -> bool:
    """用 nvidia-smi 判断本机是否有可用 NVIDIA GPU（不依赖 paddle 是否为 GPU 版）。"""
    try:
        import subprocess
        r = subprocess.run(
            ['nvidia-smi', '-L'],
            capture_output=True, text=True, encoding='utf-8', errors='replace',
            timeout=5,
        )
        return r.returncode == 0 and 'GPU' in (r.stdout or '')
    except Exception:
        return False


def _detect_paddle_device():
    """优先 GPU；可用环境变量 INVOICE_TOOL_DEVICE=gpu:0 / cpu 强制指定。"""
    forced = (os.environ.get('INVOICE_TOOL_DEVICE') or '').strip().lower()
    if forced in ('cpu', 'gpu', 'gpu:0') or forced.startswith('gpu:'):
        if forced == 'gpu':
            return 'gpu:0'
        return forced

    _ensure_torch_before_paddle()
    try:
        import paddle
        if paddle.device.is_compiled_with_cuda():
            try:
                if paddle.device.cuda.device_count() > 0:
                    return 'gpu:0'
            except Exception:
                pass
        # 已编译 CUDA 但 device_count 异常时，若本机有独显仍尝试 GPU
        if paddle.device.is_compiled_with_cuda() and _nvidia_gpu_visible():
            return 'gpu:0'
    except Exception:
        pass

    # 有 NVIDIA 显卡但装了 CPU 版 paddle 时给出明确提示
    if _nvidia_gpu_visible():
        sys.stderr.write(
            "  [检测] 发现 NVIDIA GPU，但当前 Paddle 为 CPU 版，将使用 CPU。\n"
            "  [提示] 请安装 GPU 版: 见 requirements-gpu.txt 或 README\n"
        )
        sys.stderr.flush()
    return 'cpu'


def _normalize_match_text(text: str) -> str:
    """统一全角标点，便于跨 OCR 引擎匹配。"""
    if not text:
        return ''
    return (
        text.replace('：', ':')
        .replace('（', '(')
        .replace('）', ')')
        .replace('￥', '¥')
    )


def _parse_ocr_result(result) -> list:
    """兼容 PaddleOCR 2.x / 3.x 多种 predict 返回格式。"""
    texts = []
    if not result:
        return texts

    items = result if isinstance(result, (list, tuple)) else [result]
    for item in items:
        chunk = []
        if isinstance(item, dict):
            chunk = item.get('rec_texts') or []
        elif hasattr(item, 'get'):
            chunk = item.get('rec_texts') or []
        if not chunk and hasattr(item, 'json'):
            payload = item.json() if callable(item.json) else item.json
            if isinstance(payload, dict):
                res = payload.get('res', payload)
                if isinstance(res, dict):
                    chunk = res.get('rec_texts') or []
        if chunk:
            texts.extend(chunk)

    # PaddleOCR 2.x: [[[box], (text, score)], ...]
    if not texts and items and isinstance(items[0], (list, tuple)):
        for page in items:
            if not page:
                continue
            for line in page:
                if isinstance(line, (list, tuple)) and len(line) >= 2:
                    text_part = line[1]
                    if isinstance(text_part, (list, tuple)) and text_part:
                        texts.append(str(text_part[0]))
                    elif isinstance(text_part, str):
                        texts.append(text_part)
    return texts


def _run_ocr_on_image(ocr, image_path: str) -> str:
    """对单张图片执行 OCR，兼容 predict / ocr 两种 API。"""
    if hasattr(ocr, 'predict'):
        try:
            texts = _parse_ocr_result(ocr.predict(image_path))
            if texts:
                return '\n'.join(texts)
        except Exception as e:
            sys.stderr.write(f"  [警告] OCR predict 失败: {e}\n")
            sys.stderr.flush()

    if hasattr(ocr, 'ocr'):
        try:
            legacy = ocr.ocr(image_path, cls=False)
            texts = _parse_ocr_result(legacy)
            if texts:
                return '\n'.join(texts)
        except Exception as e:
            sys.stderr.write(f"  [警告] OCR ocr() 失败: {e}\n")
            sys.stderr.flush()
    return ''


def _render_page_pixmap(page, target_max_side: int = None, dpi: int = None):
    """按当前识别模式渲染 PDF 页面（不超过 PaddleOCR max_side_limit≈4000）。"""
    device = _detect_paddle_device()
    prof = _current_ocr_profile()
    is_cpu = (device == 'cpu')
    if dpi is None:
        dpi = prof['dpi_cpu'] if is_cpu else prof['dpi_gpu']
    if target_max_side is None:
        target_max_side = prof['max_side_cpu'] if is_cpu else prof['max_side_gpu']
    rect = page.rect
    if rect.width > 0 and rect.height > 0:
        max_side = max(rect.width, rect.height) / 72.0 * dpi
        if max_side > target_max_side:
            dpi = max(120 if is_cpu else 160, int(dpi * target_max_side / max_side))
    return page.get_pixmap(dpi=dpi, alpha=False)


def _ocr_page_text(ocr, fitz_page, page_label: str = '') -> str:
    """对 PDF 单页执行 OCR；精度/平衡模式下字段不全时可提高清晰度重试。"""
    hint = f' ({page_label})' if page_label else ''
    prof = _current_ocr_profile()
    _emit_progress(f"  [OCR] 本页图像识别中{hint}（{prof['label']}模式）...")
    pix = _render_page_pixmap(fitz_page)
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    tmp.close()
    try:
        pix.save(tmp.name)
        text = _run_ocr_on_image(ocr, tmp.name)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    allow_retry = bool(prof.get('retry_incomplete'))
    if allow_retry and text and _classify_page(text, source='ocr') == 'invoice':
        probe = InvoiceInfo()
        _extract_fields(text, probe)
        need_retry = (
            not probe.invoice_number
            or not probe.total_with_tax
            or (not probe.buyer_name and not probe.seller_name)
        )
        if need_retry and _detect_paddle_device().startswith('gpu'):
            retry_dpi = int(prof.get('retry_dpi_gpu') or min(320, prof['dpi_gpu'] + 60))
            retry_side = min(3900, int(prof['max_side_gpu']))
            _emit_progress(f"  [OCR] 字段不完整，提高清晰度重试{hint}...")
            pix2 = _render_page_pixmap(fitz_page, target_max_side=retry_side, dpi=retry_dpi)
            tmp2 = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            tmp2.close()
            try:
                pix2.save(tmp2.name)
                text2 = _run_ocr_on_image(ocr, tmp2.name)
                if text2 and len(text2) >= len(text) * 0.8:
                    probe2 = InvoiceInfo()
                    _extract_fields(text2, probe2)
                    score = lambda p: sum(bool(x) for x in (
                        p.invoice_number, p.invoice_date, p.buyer_name,
                        p.seller_name, p.total_with_tax,
                    ))
                    if score(probe2) >= score(probe):
                        text = text2
            finally:
                try:
                    os.unlink(tmp2.name)
                except OSError:
                    pass
    return text or ''


def _extract_receipt_number(text: str) -> str:
    """提取财政/医疗电子票据号码（票据号码），兼容 OCR 误识别。
    注意：不得误匹配增值税发票的「发票号码」。"""
    norm = _normalize_match_text(text)
    compact = re.sub(r'\s+', '', norm)

    patterns = [
        r'票据号码[：:.]\s*(\d{6,30})',
        r'票据号[码马][：:.]\s*(\d{6,30})',
        r'柔据号[码马][：:.]\s*(\d{6,30})',
        r'票[据櫃][号码码]+[：:.]\s*(\d{6,30})',
        # 避免匹配「发票号码」：号码前一字不能是「票」
        r'(?<![发票票代代码马])号码[：:.](\d{6,30})',
        r'(?<![发票票代代码马])号码(\d{8,12})',
    ]
    for pattern in patterns:
        m = re.search(pattern, compact)
        if m:
            return m.group(1)

    m = re.search(r'票据号[码马][：:OIlB.]*([0-9OIlB]{6,30})', compact)
    if m:
        raw = m.group(1)
        trans = str.maketrans({'O': '0', 'o': '0', 'I': '1', 'l': '1', 'B': '8'})
        num = raw.translate(trans)
        num = re.sub(r'[^0-9]', '', num)
        if len(num) >= 6:
            return num
    return ''


def _normalize_ocr_text(text: str) -> str:
    """修正常见 OCR 误识别，便于后续字段提取。"""
    if not text:
        return ''
    fixes = [
        ('开东日期', '开票日期'),
        ('开票月期', '开票日期'),
        ('开累日期', '开票日期'),
        ('开乘日期', '开票日期'),
        ('开系日期', '开票日期'),
        ('开票日斯', '开票日期'),
        ('收款单住', '收款单位'),
        ('收杖单位', '收款单位'),
        ('收狄单位', '收款单位'),
        ('收欣单位', '收款单位'),
        ('收状单位', '收款单位'),
        ('文款人、', '文款人：'),
        ('文款人,', '文款人：'),
        ('文秋人、', '文秋人：'),
        ('交款人—', '交款人：'),
        ('交款人-', '交款人：'),
        ('此诊日期', '就诊日期'),
        ('柔据号码', '票据号码'),
        ('条据号码', '票据号码'),
        ('集据号码', '票据号码'),
        ('票根号码', '票据号码'),
        ('金须合计', '金额合计'),
        ('金合计', '金额合计'),
        ('金颠合计', '金额合计'),
        ('个人见金支付', '个人现金支付'),
        ('个人见盒支付', '个人现金支付'),
        ('个人现金支什', '个人现金支付'),
        ('价税台计', '价税合计'),
        ('价税合汁', '价税合计'),
    ]
    for old, new in fixes:
        text = text.replace(old, new)
    # 仅修正「交款人／收款人」后的全角点号，避免误伤姓名中的间隔号
    text = re.sub(
        r'([交文又][款此秋北止使武达]?人|[收收款]款?单位)\s*[．·、]\s*',
        r'\1：',
        text,
    )
    text = re.sub(r'([人位码期])\s*[：:]\s*', r'\1：', text)
    return text


def _looks_like_person_name(name: str) -> bool:
    """判断字符串是否像中文人名。"""
    if not name or len(name) < 2 or len(name) > 8:
        return False
    if re.search(r'社会|信用|校验|票据|代码|开票|统一|项目|卫生院|医院|公司|门诊|体检|集团|有限', name):
        return False
    if re.fullmatch(r'[\d*~.]+', name):
        return False
    return bool(re.search(r'[\u4e00-\u9fff]', name))


def _parse_compact_date(raw: str) -> str:
    """将 YYYYMMDD / 缺位 YYYMMDD 转为 YYYY-MM-DD。"""
    digits = re.sub(r'\D', '', raw or '')
    if len(digits) == 7:
        if digits.startswith('206'):
            digits = '2026' + digits[3:]
        elif digits.startswith('205'):
            digits = '2025' + digits[3:]
        elif digits.startswith('204'):
            digits = '2024' + digits[3:]
        elif digits.startswith('20'):
            digits = digits[:4] + '0' + digits[4:]
    if len(digits) == 8:
        return f'{digits[:4]}-{digits[4:6]}-{digits[6:8]}'
    return ''


def _extract_invoice_date(text: str) -> str:
    """提取开票日期，兼容 OCR 误识别及就诊日期回退。"""
    norm = _normalize_match_text(_normalize_ocr_text(text))

    m = re.search(r'开[票东累乘系]日期[：:.]?\s*(\d{4})[年/-](\d{1,2})[月/-](\d{1,2})', norm)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"

    m = re.search(r'开[票东累乘系]日期[：:.]?\s*(\d{7,8})', norm)
    if m:
        parsed = _parse_compact_date(m.group(1))
        if parsed:
            return parsed

    m = re.search(r'开[票东累乘系]日期[：:.]?\s*\n\s*(\d{8})', norm)
    if m:
        parsed = _parse_compact_date(m.group(1))
        if parsed:
            return parsed

    # OCR 把「开票日期：20260124」识别成中间夹杂空格/符号
    m = re.search(r'开[票东累乘系]日期[：:.]?\s*([0-9OIlB\s]{7,12})', norm)
    if m:
        digits = re.sub(r'[^0-9]', '', m.group(1).translate(str.maketrans('OIlB', '0118')))
        parsed = _parse_compact_date(digits)
        if parsed:
            return parsed

    for pattern in (r'就诊日期[：:.]?\s*(\d{8})', r'就诊日期[：:.]?\s*(\d{4})(\d{2})(\d{2})'):
        m = re.search(pattern, norm)
        if m:
            if m.lastindex == 1:
                parsed = _parse_compact_date(m.group(1))
                if parsed:
                    return parsed
            else:
                return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return ''


def _extract_payer_name(text: str) -> str:
    """提取财政票据交款人，兼容 OCR 误识别及姓名换行。"""
    norm = _normalize_match_text(_normalize_ocr_text(text))
    cn_name = r'[\u4e00-\u9fff·]{2,8}'
    punct = r'[：:.\s．·、]*'

    inline_patterns = [
        rf'[交文又][款此秋北止使武达]人{punct}({cn_name})',
        rf'文人{punct}({cn_name})',
        rf'又北人{punct}({cn_name})',
        rf'交[北武]人{punct}({cn_name})',
        rf'交款人{punct}({cn_name})',
        rf'文此人{punct}({cn_name})',
        rf'交此人{punct}({cn_name})',
        rf'[交文又][款此秋北止使武达]人({cn_name})',
        rf'文达人({cn_name})',
    ]
    for pattern in inline_patterns:
        for m in re.finditer(pattern, norm):
            name = _clean_name(m.group(1).split('统一')[0].split('校验')[0].split('开票')[0])
            if _looks_like_person_name(name):
                return name

    multiline_patterns = [
        rf'久止人\s*\n\s*({cn_name})',
        rf'(?:[交文又][款此秋北止使武达]人|文人|文此人)\s*\n\s*({cn_name})',
    ]
    for pattern in multiline_patterns:
        m = re.search(pattern, norm)
        if m:
            name = _clean_name(m.group(1))
            if _looks_like_person_name(name):
                return name

    # 性别行前常见姓名残留（如「张三\n性别:男」）
    m = re.search(rf'({cn_name})\s*\n\s*性别', norm)
    if m and _looks_like_person_name(m.group(1)):
        return m.group(1)
    return ''


def _extract_receipt_seller(text: str) -> str:
    """提取财政票据收款单位，兼容 OCR 误识别。"""
    norm = _normalize_ocr_text(text)
    patterns = [
        r'收款单位[：:]\s*(.+?)(?:\n|复核人|核人|开票人|$)',
        r'收[款杖狄状]?单?[位住][：:]\s*(.+?)(?:\n|复核人|核人|$)',
        r'收款单位\s*\n\s*([^\n]+)',
        r'收[款状]?单位\s*\n\s*([^\n]+)',
        r'收款单位\s*([^\n]{4,30})',
    ]
    for pattern in patterns:
        m = re.search(pattern, norm, re.DOTALL)
        if m:
            name = _clean_name(m.group(1).split('\n')[0])
            if name and (
                re.search(r'卫生院|医院|中心|诊所|门诊部|卫生室', name)
                or (len(name) >= 4 and not re.search(r'校验|票据|代码|支付|合计', name))
            ):
                return name

    m = re.search(r'备注[：:][^\n]*?([\u4e00-\u9fff]{2,24}(?:卫生院|医院|中心|诊所))', norm)
    if m:
        return _clean_name(m.group(1))
    # 页脚机构名兜底
    m = re.search(r'([\u4e00-\u9fff]{2,16}(?:县|市|区)[\u4e00-\u9fff]{0,12}(?:卫生院|医院|中心))', norm)
    if m:
        return _clean_name(m.group(1))
    return ''


def _extract_invoice_number(text: str) -> str:
    """提取发票号码，兼容 OCR 误识别（空格、全角冒号等）。"""
    norm = _normalize_match_text(text)
    compact = re.sub(r'\s+', '', norm)

    patterns = [
        r'发票号码[：:]\s*(\d{8,30})',
        r'发票号码[：:]*([0-9]{8,30})',
        r'发票号码\s*[:：]?\s*(\d{8,30})',
        # 部分 PDF 排版：号码在「发票号码」标签上一行
        r'(\d{8,30})\s*发票号码[：:]?',
    ]
    for pattern in patterns:
        m = re.search(pattern, compact)
        if m:
            return m.group(1)

    # 独立长数字行紧邻「发票号码」标签（pdfplumber 换行排版）
    m = re.search(r'(\d{17,30})\s*\n\s*发票号码', norm)
    if m:
        return m.group(1)

    # OCR 可能把数字 0 识别成 O
    m = re.search(r'发票号码[：:OIlB]*([0-9OIlB]{8,30})', compact)
    if m:
        raw = m.group(1)
        trans = str.maketrans({'O': '0', 'o': '0', 'I': '1', 'l': '1', 'B': '8'})
        num = raw.translate(trans)
        num = re.sub(r'[^0-9]', '', num)
        if len(num) >= 8:
            return num

    return _extract_receipt_number(text)


def _get_ocr():
    """获取 PaddleOCR 实例（单例），自动处理中文路径、缓存损坏等问题"""
    global _ocr_instance, _ocr_failed, _ocr_instance_mode

    if _ocr_instance is not None:
        return _ocr_instance
    if _ocr_failed:
        raise RuntimeError("PaddleOCR 不可用（已放弃），图片型 PDF 将无法识别")

    import shutil

    # ── 检测中文用户名 ──
    _original_expanduser = os.path.expanduser
    user_home = _original_expanduser('~')
    has_unicode_path = any(ord(c) > 127 for c in user_home)

    if has_unicode_path:
        # 选可用盘符，优先 D 盘
        drives = ['D:', 'E:', 'F:', 'C:']
        safe_home = 'C:\\paddle_ocr_cache'
        for d in drives:
            if os.path.exists(d + '\\'):
                safe_home = os.path.join(d + '\\', 'paddle_ocr_cache')
                break
        os.makedirs(safe_home, exist_ok=True)

        # ── 核心：拦截 os.path.expanduser，骗过 PaddleX ──
        def _patched_expanduser(p):
            if isinstance(p, str) and p.startswith('~'):
                return p.replace('~', safe_home, 1)
            return _original_expanduser(p)
        os.path.expanduser = _patched_expanduser

        sys.stderr.write(f"  [检测] 用户名含中文，模型缓存重定向到: {safe_home}\n")
        sys.stderr.flush()

    # ── 跳过连通性检查 ──
    os.environ.setdefault('PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK', 'True')

    # 必须在 detect/import paddle 之前预加载 torch，避免 Windows DLL 冲突
    _ensure_torch_before_paddle()
    device = _detect_paddle_device()
    prof = _current_ocr_profile()
    sys.stderr.write(f"  [检测] Paddle 运行设备: {device}，识别模式: {prof['label']}\n")
    sys.stderr.flush()

    ocr_kwargs = dict(
        lang='ch',
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
    )

    def _ocr_init_kwargs(dev):
        """按模式选择 mobile / server 模型。"""
        kw = dict(ocr_kwargs)
        want_server = bool(prof.get('use_server_model')) and str(dev).startswith('gpu')
        if want_server:
            kw.update(
                text_detection_model_name='PP-OCRv4_server_det',
                text_recognition_model_name='PP-OCRv4_server_rec',
            )
        else:
            kw.update(
                text_detection_model_name='PP-OCRv4_mobile_det',
                text_recognition_model_name='PP-OCRv4_mobile_rec',
            )
        return kw

    max_retries = 3
    last_error = None

    for attempt in range(max_retries):
        # 优先 GPU，失败后强制 CPU（其他电脑无 GPU 或 GPU 驱动异常时常见）
        device_candidates = []
        if device.startswith('gpu'):
            device_candidates.append(device)
        device_candidates.append('cpu')

        for dev in device_candidates:
            try:
                sys.stderr.write(
                    f"  [初始化] 正在加载 PaddleOCR 模型 "
                    f"(尝试 {attempt+1}/{max_retries}, 设备 {dev}, 模式 {prof['label']})...\n"
                )
                sys.stderr.flush()
                _ensure_torch_before_paddle()
                from paddleocr import PaddleOCR
                init_kw = _ocr_init_kwargs(dev)
                try:
                    _ocr_instance = PaddleOCR(device=dev, **init_kw)
                except TypeError:
                    _ocr_instance = PaddleOCR(**init_kw)
                except Exception as model_err:
                    # GPU server 模型不可用时回退 mobile，避免整段放弃
                    if str(dev).startswith('gpu') and 'server_' in str(init_kw):
                        sys.stderr.write(
                            f"  [提示] server 模型加载失败，回退 mobile: {str(model_err)[:120]}\n"
                        )
                        sys.stderr.flush()
                        init_kw = dict(ocr_kwargs)
                        init_kw.update(
                            text_detection_model_name='PP-OCRv4_mobile_det',
                            text_recognition_model_name='PP-OCRv4_mobile_rec',
                        )
                        try:
                            _ocr_instance = PaddleOCR(device=dev, **init_kw)
                        except TypeError:
                            _ocr_instance = PaddleOCR(**init_kw)
                    else:
                        raise
                _ocr_instance_mode = _ocr_quality_mode
                sys.stderr.write(f"  [初始化] PaddleOCR 模型加载完成 (设备 {dev})\n")
                sys.stderr.flush()
                tip = (
                    f"  [提示] OCR 引擎就绪（{prof['label']}模式"
                    + (" / GPU" if str(dev).startswith('gpu') else " / CPU")
                    + "），即将逐页识别。"
                )
                _emit_progress(tip)
                if has_unicode_path:
                    os.path.expanduser = _original_expanduser
                return _ocr_instance
            except Exception as e:
                last_error = e
                err_msg = str(e)
                sys.stderr.write(f"  [错误] PaddleOCR 加载失败 ({dev}): {err_msg[:200]}\n")
                sys.stderr.flush()
                _ocr_instance = None

        err_msg = str(last_error or '')
        is_cache_error = any(pattern in err_msg for pattern in [
            'parse_error', 'parse error', 'parse an empty',
            'empty input', 'corrupt', 'unexpected end of file',
        ])

        if is_cache_error and attempt < max_retries - 1:
            sys.stderr.write("  [修复] 正在清理缓存...\n")
            sys.stderr.flush()
            cache_dirs = list(get_paddle_cache_dirs())
            for cache_path in cache_dirs:
                if os.path.exists(cache_path):
                    try:
                        shutil.rmtree(cache_path)
                        sys.stderr.write(f"  [清理] 已删除: {cache_path}\n")
                    except Exception:
                        sys.stderr.write(f"  [警告] 无法删除: {cache_path}\n")
            sys.stderr.flush()
        else:
            break

    # ── 重试耗尽 ──
    if has_unicode_path:
        os.path.expanduser = _original_expanduser
    _ocr_failed = True
    msg = (
        f"\n  {'='*50}\n"
        f"  PaddleOCR 初始化失败（重试 {max_retries} 次）\n"
        f"  错误: {str(last_error)[:300]}\n"
        f"  {'='*50}\n"
        f"  图片型 PDF 将无法识别，文本型 PDF 不受影响。\n"
        f"\n"
        f"  请尝试：\n"
        f"  1. 删除文件夹 {os.path.join(safe_home if has_unicode_path else user_home, '.paddlex')}\n"
        f"  2. 确保网络畅通，重新运行本程序\n"
        f"  {'='*50}\n"
    )
    sys.stderr.write(msg)
    sys.stderr.flush()
    raise RuntimeError(f"PaddleOCR 不可用: {last_error}")


class InvoiceInfo:
    """单张发票的信息"""
    def __init__(self):
        self.file_name = ""
        self.source_folder = ""       # 来源文件夹名（批量模式用）
        self.invoice_number = ""
        self.invoice_date = ""
        self.buyer_name = ""
        self.seller_name = ""
        self.subtotal = ""
        self.total_with_tax = ""
        self.phone_number = ""
        self.billing_period = ""
        self.extraction_method = ""
        self.invoice_type = ""        # 自动识别：vat / rail / fiscal / unknown


# 发票类型常量（自动识别，无需用户选择）
INVOICE_TYPE_VAT = 'vat'          # 增值税电子发票（普通/专用等）
INVOICE_TYPE_RAIL = 'rail'        # 铁路电子客票
INVOICE_TYPE_FISCAL = 'fiscal'    # 财政/医疗电子票据
INVOICE_TYPE_UNKNOWN = 'unknown'


def detect_invoice_type(text: str) -> str:
    """根据页面文本自动判断发票类型，供后续分流提取。"""
    if not text:
        return INVOICE_TYPE_UNKNOWN
    compact = re.sub(r'\s+', '', _normalize_match_text(_normalize_ocr_text(text)))

    # 1) 铁路电子客票
    rail_hits = sum([
        bool(re.search(r'铁路电子客票', compact)),
        bool(re.search(r'电子客票号', compact)),
        bool(re.search(r'票价[：:].*[¥￥]|票价[¥￥]', compact)),
        bool(re.search(r'12306|中国铁路|国家铁路', compact)),
        bool(re.search(r'席别|车次|始发|到达站|出发站', compact)),
    ])
    if rail_hits >= 2 or (re.search(r'铁路电子客票|电子客票号', compact) and re.search(r'发票号码|电子发票', compact)):
        return INVOICE_TYPE_RAIL

    # 2) 增值税电子发票（优先于财政票，避免「小写/号码」误判）
    vat_hits = sum([
        bool(re.search(r'价税合计', compact)),
        bool(re.search(r'购买方信息|销售方信息', compact)),
        bool(re.search(r'增值税|普通发票|专用发票', compact)),
        bool(re.search(r'纳税人识别号', compact)),
        bool(re.search(r'发票号码', compact) and re.search(r'电子发票', compact)),
    ])
    if vat_hits >= 2:
        return INVOICE_TYPE_VAT
    if re.search(r'价税合计', compact) and re.search(r'购买方|销售方|发票号码', compact):
        return INVOICE_TYPE_VAT

    # 3) 财政/医疗电子票据
    if re.search(r'票据号码|柔据号码|财政电子票据|医疗.*收费票据|门诊收费票据|浙江省医疗|卫生院', compact):
        if re.search(r'交款人|收款单位|金额合计|交此人|文款人|文此人|小写', compact) or _extract_receipt_number(text):
            return INVOICE_TYPE_FISCAL
    if _is_fiscal_receipt_page(text):
        return INVOICE_TYPE_FISCAL

    if re.search(r'电子发票|增值税电子', compact):
        return INVOICE_TYPE_VAT

    return INVOICE_TYPE_UNKNOWN


def _clean_amount(raw: str) -> str:
    """规范化金额字符串，去掉逗号、空格与残缺小数点。"""
    if not raw:
        return ''
    s = str(raw).replace(',', '').replace(' ', '').replace('￥', '').replace('¥', '').strip()
    m = re.search(r'(\d+\.\d{2})', s)
    if m:
        return m.group(1)
    m = re.search(r'^(\d+)\.(\d)$', s)
    if m:
        return f"{m.group(1)}.{m.group(2)}0"
    m = re.search(r'^(\d+)\.$', s)
    if m:
        return ''
    m = re.search(r'^(\d+)(?:\.(\d+))?$', s)
    if m:
        if m.group(2) is None:
            return m.group(1)
        return f"{m.group(1)}.{m.group(2)[:2]}"
    return ''


def _extract_amount_from_text(text: str) -> str:
    """从多种金额写法中提取，兼容 OCR 空格（如 176. 00）。"""
    patterns = [
        r'价税合计.*?[（(]小写[）)]\s*[¥￥]?\s*(\d+\s*\.\s*\d{2})',
        r'[（(]小写[）)]\s*[¥￥]?\s*(\d+\s*\.\s*\d{2})',
        r'金额合计.*?[（(]小写[）)]\s*[¥￥]?\s*(\d+\s*\.\s*\d{2})',
        r'个人[见现]?金支付[：:]\s*[¥￥]?\s*(\d+\s*\.\s*\d{2})',
        r'票价[：:]\s*[¥￥]?\s*(\d+\s*\.\s*\d{2})',
        r'[（(]小写[）)]\s*[¥￥]?\s*([\d,]+\.?\d*)',
        r'金额合计.*?[（(]小写[）)]\s*([\d,]+\.?\d*)',
        r'个人[见现]?金支付[：:]\s*([\d,]+\.?\d*)',
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.DOTALL)
        if m:
            amt = _clean_amount(m.group(1))
            if amt and '.' in amt:
                return amt
    for pattern in patterns:
        m = re.search(pattern, text, re.DOTALL)
        if m:
            amt = _clean_amount(m.group(1))
            if amt:
                return amt
    return ''


def _clean_name(name):
    """清理名称中混入的 PDF 排版碎片文字"""
    name = re.sub(r'\s+', '', name)
    name = re.sub(r'[买卖售方信]+$', '', name)
    for sep in ['买售方', '买售', '买']:
        idx = name.find(sep)
        if idx > 0:
            name = name[:idx]
            break
    return name.strip()


def _is_non_invoice_page(text: str) -> bool:
    """判断页面是否明确不是电子发票（报销单、明细表等），应跳过"""
    # 去除所有空白，方便匹配
    compact = re.sub(r'\s+', '', text)

    # 1. "报销单" 单独出现（如 "通讯费报销单"），但排除发票中 "报销" 出现的情况
    if re.search(r'报销单', compact):
        # 如果同时有发票号码 + 开票日期，可能是发票页里恰好出现了"报销"字样
        has_invoice_markers = (
            re.search(r'发票号码[：:]\s*\d+', compact)
            and re.search(r'开票日期', compact)
        )
        if not has_invoice_markers:
            return True

    # 2. "明细表" 单独出现（如 "通讯费报销明细表"、"银行卡明细表"）
    if re.search(r'明细表', compact):
        has_invoice_markers = (
            re.search(r'发票号码[：:]\s*\d+', compact)
            and re.search(r'开票日期', compact)
        )
        if not has_invoice_markers:
            return True

    # 3. "银行卡明细表" —— 绝对非发票
    if re.search(r'银行卡', compact) and re.search(r'明细', compact):
        return True

    return False


def _is_fiscal_receipt_page(text: str) -> bool:
    """判断是否为财政/医疗门诊电子票据（票据号码 + 开票日期/金额）。"""
    compact = re.sub(r'\s+', '', _normalize_match_text(text))
    has_receipt_number = bool(_extract_receipt_number(text))
    if not has_receipt_number:
        return False

    has_receipt_title = bool(re.search(
        r'医疗.*收费票据|财政电子票据|门诊收费票据|浙江省医疗',
        compact,
    ))
    has_payer = bool(re.search(r'交款人|交此人|文款人|文秋人', compact))
    has_amount = bool(re.search(r'金额合计|小写', compact))
    has_date = bool(re.search(r'开票日期', compact))

    if has_receipt_title:
        return True
    if has_receipt_number and has_date and (has_payer or has_amount):
        return True
    return False


def _is_invoice_page(text: str, source: str = 'text') -> bool:
    """判断页面是否为电子发票（必须有发票号码等核心特征）"""
    if _is_fiscal_receipt_page(text):
        return True

    compact = re.sub(r'\s+', '', _normalize_match_text(text))

    has_invoice_number = bool(_extract_invoice_number(text))

    has_invoice_date = bool(re.search(r'开票日期[：:]', compact))
    has_einvoice_title = bool(re.search(r'电子发票|增值税电子', compact))
    has_buyer_seller = (
        bool(re.search(r'购买方', compact))
        and bool(re.search(r'销售方', compact))
    )
    has_total = bool(re.search(r'价税合计', compact))
    has_rail = bool(re.search(r'铁路电子客票|电子客票号|票价', compact))

    if not has_invoice_number:
        return False

    # 铁路电子客票：有发票号码 + 铁路特征即可
    if has_rail and (has_einvoice_title or has_invoice_date or re.search(r'电子客票', compact)):
        return True

    auxiliary_score = sum([has_invoice_date, has_einvoice_title, has_buyer_seller, has_total])
    if auxiliary_score >= 1:
        return True

    # OCR 文本常缺辅助字段，但有发票号码且非明确非发票页时仍视为发票
    if source == 'ocr' and not _is_non_invoice_page(text):
        return True

    return False


def _classify_page(text: str, source: str = 'text') -> str:
    """
    页面分类：返回 'invoice' / 'non_invoice' / 'uncertain'
    - invoice: 确认是电子发票，可以提取
    - non_invoice: 确认不是发票（报销单、明细表等），跳过
    - uncertain: 无法确定，保守跳过（避免误提取）
    """
    if _is_non_invoice_page(text):
        return 'non_invoice'

    if _is_invoice_page(text, source=source):
        return 'invoice'

    return 'uncertain'


def _extract_buyer_seller_vat(text: str, info: InvoiceInfo):
    """增值税电子发票：购买方 / 销售方。"""
    company_keywords = ['公司', '移动', '电信', '联通', '集团', '有限', '股份']

    def _is_company(name: str) -> bool:
        return any(kw in name for kw in company_keywords)

    def _take_name(raw: str) -> str:
        name = _clean_name((raw or '').split('\n')[0].split('统一')[0])
        if not name or len(name) <= 1 or name.startswith('统一') or '信用代码' in name:
            return ''
        return name

    # 同行双名称：名称：买家 名称：卖家（pdfplumber 常见）
    m = re.search(
        r'名称[：:]\s*([^\n]+?)\s+名称[：:]\s*([^\n]+?)(?:\s*统一|\s*$|\n)',
        text,
    )
    if m:
        n1, n2 = _take_name(m.group(1)), _take_name(m.group(2))
        if n1 and n2:
            if _is_company(n1) and not _is_company(n2):
                info.seller_name, info.buyer_name = n1, n2
            elif _is_company(n2) and not _is_company(n1):
                info.buyer_name, info.seller_name = n1, n2
            else:
                info.buyer_name, info.seller_name = n1, n2

    if not info.buyer_name:
        m = re.search(r'购买方名称[：:]\s*([^\n]+)', text)
        if m:
            info.buyer_name = _take_name(m.group(1))
    if not info.seller_name:
        m = re.search(r'销售方名称[：:]\s*([^\n]+)', text)
        if m:
            info.seller_name = _take_name(m.group(1))

    if info.buyer_name and info.seller_name:
        return

    is_pdfplumber_layout = bool(re.search(r'[购销].{0,5}名\s*称[：:]', text))
    is_ocr_layout = bool(re.search(r'购买方信息|销售方信息', text))

    if is_pdfplumber_layout and not is_ocr_layout:
        if not info.buyer_name:
            m = re.search(r'购.*?名称[：:]\s*([^\n]+?)(?:\s+销|\s+名称[：:]|\s*统一|$)', text, re.DOTALL)
            if m:
                info.buyer_name = _take_name(m.group(1))
        if not info.seller_name:
            m = re.search(r'销.*?名称[：:]\s*([^\n]+?)(?:\s+名称[：:]|\s*统一|$)', text, re.DOTALL)
            if m:
                info.seller_name = _take_name(m.group(1))
        return

    buyer_block = re.search(r'购买方信息(.*?)(?:销售方信息|$)', text, re.DOTALL)
    seller_block = re.search(r'销售方信息(.*?)(?:项目名称|合计|价税合计|备注|$)', text, re.DOTALL)
    if buyer_block and not info.buyer_name:
        m = re.search(r'名称[：:]\s*([^\n]+)', buyer_block.group(1))
        if m:
            info.buyer_name = _take_name(m.group(1))
    if seller_block and not info.seller_name:
        names = re.findall(r'名称[：:]\s*([^\n]+)', seller_block.group(1))
        cleaned = [n for n in (_take_name(x) for x in names) if n]
        company_names = [n for n in cleaned if _is_company(n)]
        if company_names:
            info.seller_name = company_names[0]
        elif cleaned:
            for n in cleaned:
                if n != info.buyer_name:
                    info.seller_name = n
                    break
            if not info.seller_name:
                info.seller_name = cleaned[0]

    if info.buyer_name and info.seller_name:
        return

    all_name_matches = list(re.finditer(r'名称[：:]\s*([^\n]+)', text))
    buyer_candidates, seller_candidates = [], []
    for match in all_name_matches:
        name = _take_name(match.group(1))
        if not name:
            continue
        if _is_company(name):
            seller_candidates.append(name)
        else:
            buyer_candidates.append(name)

    if not info.seller_name and seller_candidates:
        info.seller_name = seller_candidates[0]
    if not info.buyer_name and buyer_candidates:
        info.buyer_name = buyer_candidates[0]

    if not info.seller_name:
        m = re.search(r'销.*?名称[：:]\s*([^\n]+?)(?:\s+名称[：:]|\s*统一|$)', text, re.DOTALL)
        if m:
            info.seller_name = _take_name(m.group(1))
    if not info.buyer_name:
        m = re.search(r'购.*?名称[：:]\s*([^\n]+?)(?:\s+销|\s+名称[：:]|\s*统一|$)', text, re.DOTALL)
        if m:
            info.buyer_name = _take_name(m.group(1))
    if not info.buyer_name:
        m = re.search(r'购买方.*?名称[：:]\s*([^\n]+)', text, re.DOTALL)
        if m:
            info.buyer_name = _take_name(m.group(1))



def _extract_rail_fields(text: str, info: InvoiceInfo):
    """铁路电子客票专用字段。"""
    m = re.search(r'购买方名称[：:]\s*([^\n]+)', text)
    if m:
        name = _clean_name(m.group(1).split('统一')[0])
        if name and '信用代码' not in name:
            info.buyer_name = name

    # 无企业购方时，身份证号下一行常为乘客姓名
    if not info.buyer_name:
        m2 = re.search(
            r'\d{6}\d{4}\*{4}\d{4}\s*\n\s*([\u4e00-\u9fff·]{2,8})',
            text,
        )
        if m2 and _looks_like_person_name(m2.group(1)):
            info.buyer_name = m2.group(1)

    m = re.search(r'票价[：:]\s*[¥￥]?\s*([\d,]+\.?\d*)', text)
    if m:
        amt = _clean_amount(m.group(1))
        if amt:
            info.total_with_tax = amt
            if not info.subtotal:
                info.subtotal = amt

    # 铁路票通常不印销售方全称
    if not info.seller_name:
        if re.search(r'中国铁路|国家铁路|12306|铁路电子客票|中国国家铁路', text):
            info.seller_name = '中国国家铁路集团有限公司'


def _extract_fiscal_fields(text: str, info: InvoiceInfo):
    """财政 / 医疗电子票据专用字段。"""
    if not info.buyer_name:
        info.buyer_name = _extract_payer_name(text)
    if not info.seller_name:
        info.seller_name = _extract_receipt_seller(text)

    if not info.total_with_tax:
        info.total_with_tax = _extract_amount_from_text(text)

    # 交款人标签被 OCR 拆坏时，尝试「校验码」与「开票日期」之间的短中文名
    if not info.buyer_name:
        m = re.search(
            r'校验码[：:]?\s*\d+\s*\n\s*([\u4e00-\u9fff·]{2,8})\s*\n\s*开',
            text,
        )
        if m and _looks_like_person_name(m.group(1)):
            info.buyer_name = m.group(1)


def _extract_amount_vat(text: str, info: InvoiceInfo):
    """增值税发票金额字段。"""
    m = re.search(r'合\s*计\s*[¥￥]\s*([\d,]+\.?\d*)', text)
    if m:
        info.subtotal = _clean_amount(m.group(1)) or m.group(1).replace(',', '')

    amt = _extract_amount_from_text(text)
    if amt:
        info.total_with_tax = amt
    if not info.total_with_tax:
        m = re.search(r'价税合计.*?[¥￥]\s*([\d,]+\.?\d*)', text, re.DOTALL)
        if m:
            info.total_with_tax = _clean_amount(m.group(1))


def _extract_phone_and_period(text: str, info: InvoiceInfo):
    """通讯类发票附加字段。"""
    m = re.search(r'电话\s*号?\s*码?\s*[：:]\s*(\d{11})', text)
    if m:
        info.phone_number = m.group(1)
    else:
        m = re.search(r'电话\s*[：:]\s*(\d{11})', text)
        if m:
            info.phone_number = m.group(1)
    if not info.phone_number:
        m = re.search(r'手机号码[：:]\s*(\d{11})', text)
        if m:
            info.phone_number = m.group(1)
    if not info.phone_number:
        m = re.search(r'业务号码[：:]\s*(\d{11})', text)
        if m:
            info.phone_number = m.group(1)

    m = re.search(r'计费[时段周期]+[：:]\s*(\d{6})', text)
    if m:
        period = m.group(1)
        info.billing_period = f"{period[:4]}-{period[4:6]}"
    if not info.billing_period:
        m = re.search(r'计费时段[：:]\s*(\d{4})年(\d{2})月', text)
        if m:
            info.billing_period = f"{m.group(1)}-{m.group(2).zfill(2)}"
    if not info.billing_period:
        m = re.search(r'计费[时段周期]+[：:]\s*(\d{4})[年/-](\d{2})', text)
        if m:
            info.billing_period = f"{m.group(1)}-{m.group(2).zfill(2)}"
    if not info.billing_period:
        m = re.search(r'账期[：:]\s*(\d{6})', text)
        if m:
            period = m.group(1)
            info.billing_period = f"{period[:4]}-{period[4:6]}"


def _extract_fields(text: str, info: InvoiceInfo):
    """自动识别发票类型并分流提取；用户无需手动选择类型。"""
    text = _normalize_ocr_text(text)
    inv_type = detect_invoice_type(text)
    info.invoice_type = inv_type

    # 通用：号码 + 日期
    info.invoice_number = _extract_invoice_number(text)
    info.invoice_date = _extract_invoice_date(text)

    if inv_type == INVOICE_TYPE_RAIL:
        _extract_rail_fields(text, info)
    elif inv_type == INVOICE_TYPE_FISCAL:
        _extract_fiscal_fields(text, info)
        # 财政票若销售方仍空，再试通用购销（少数混排）
        if not info.seller_name or not info.buyer_name:
            _extract_buyer_seller_vat(text, info)
        if not info.total_with_tax:
            _extract_amount_vat(text, info)
    else:
        # vat / unknown：走增值税通用逻辑，再用铁路/财政兜底补缺
        _extract_buyer_seller_vat(text, info)
        _extract_amount_vat(text, info)
        if not info.buyer_name or not info.total_with_tax:
            # 未判成 rail 但出现票价时补一次
            if re.search(r'票价[：:].*[¥￥]|铁路电子客票|电子客票号', text):
                _extract_rail_fields(text, info)
        if not info.buyer_name:
            info.buyer_name = _extract_payer_name(text)
        if not info.seller_name:
            info.seller_name = _extract_receipt_seller(text)
        if not info.total_with_tax:
            _extract_fiscal_fields(text, info)

    # 通讯费附加字段（仅当文本里出现时才有值）
    _extract_phone_and_period(text, info)

    # 金额最终清洗
    if info.total_with_tax:
        info.total_with_tax = _clean_amount(info.total_with_tax) or info.total_with_tax
    if info.subtotal:
        info.subtotal = _clean_amount(info.subtotal) or info.subtotal


def _build_invoice_from_text(full_text, table_text, pdf_path, source_folder, method):
    """从页面文本构建 InvoiceInfo。"""
    info = InvoiceInfo()
    info.file_name = os.path.basename(pdf_path)
    info.source_folder = source_folder
    info.extraction_method = method
    _extract_fields(full_text, info)
    if not info.subtotal:
        m = re.search(r'合\s*计\s*[¥￥]\s*([\d,]+\.?\d*)', table_text)
        if m:
            info.subtotal = _clean_amount(m.group(1)) or m.group(1).replace(',', '')
    if not info.total_with_tax:
        m = re.search(r'价税合计.*?[¥￥]\s*([\d,]+\.?\d*)', table_text, re.DOTALL)
        if m:
            info.total_with_tax = _clean_amount(m.group(1))
        if not info.total_with_tax:
            m = re.search(r'票价[：:]\s*[¥￥]?\s*([\d,]+\.?\d*)', table_text)
            if m:
                info.total_with_tax = _clean_amount(m.group(1))
    return info


def _extract_via_pdfplumber(pdf_path: str, source_folder: str = ""):
    """使用 pdfplumber 解析文本型 PDF，返回 InvoiceInfo 列表（仅电子发票页）"""
    invoices = []
    skipped_pages = 0

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                full_text = ""
                all_table_data = []

                text = page.extract_text()
                if text:
                    full_text += text + "\n"

                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        cell_texts = [cell for cell in row if cell]
                        all_table_data.extend(cell_texts)

                table_text = "\n".join(all_table_data)

                if not full_text.strip():
                    continue

                # ★ 页面分类：只提取电子发票页
                page_type = _classify_page(full_text, source='text')
                if page_type != 'invoice':
                    skipped_pages += 1
                    continue

                info = _build_invoice_from_text(
                    full_text, table_text, pdf_path, source_folder, "pdfplumber"
                )

                if info.invoice_number:
                    invoices.append(info)

        if skipped_pages > 0:
            print(
                f"  [过滤] '{os.path.basename(pdf_path)}' "
                f"跳过 {skipped_pages}/{len(pdf.pages)} 个非发票页面"
            )

    except Exception as e:
        print(f"  [错误] pdfplumber 解析 {os.path.basename(pdf_path)} 时出错: {e}")

    return invoices


def _extract_via_ocr(pdf_path: str, source_folder: str = ""):
    """使用 PaddleOCR 识别图片型 PDF，返回 InvoiceInfo 列表（仅电子发票页）"""
    invoices = []
    skipped_pages = 0

    try:
        ocr = _get_ocr()
        doc = fitz.open(pdf_path)

        for page_num in range(doc.page_count):
            page = doc[page_num]
            try:
                ocr_text = _ocr_page_text(ocr, page)
                if not ocr_text.strip():
                    continue

                page_type = _classify_page(ocr_text, source='ocr')
                if page_type != 'invoice':
                    skipped_pages += 1
                    continue

                info = _build_invoice_from_text(
                    ocr_text, ocr_text, pdf_path, source_folder, "OCR"
                )
                if info.invoice_number:
                    invoices.append(info)
            except Exception as e:
                sys.stderr.write(
                    f"  [警告] OCR 第 {page_num + 1} 页失败 "
                    f"({os.path.basename(pdf_path)}): {e}\n"
                )
                sys.stderr.flush()

        total_pages = doc.page_count
        doc.close()

        if skipped_pages > 0:
            sys.stderr.write(
                f"  [过滤] '{os.path.basename(pdf_path)}' "
                f"跳过 {skipped_pages}/{total_pages} 个非发票页面\n"
            )
            sys.stderr.flush()

    except RuntimeError:
        pass  # _get_ocr 已输出详细错误信息
    except Exception as e:
        sys.stderr.write(f"  [错误] OCR 解析 {os.path.basename(pdf_path)} 时出错: {e}\n")
        import traceback
        traceback.print_exc()

    return invoices


def _is_image_based_pdf(pdf_path: str) -> bool:
    """快速检测 PDF 是否为图片型（无实质文本层）。
    检查前3页，总字符数 < 100 → 判定为图片型。"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_chars = 0
            check_pages = min(3, len(pdf.pages))
            for i in range(check_pages):
                text = pdf.pages[i].extract_text()
                if text:
                    total_chars += len(text.strip())
            return total_chars < 100
    except Exception:
        # 无法判断时按图片型处理，走 OCR 更稳妥
        return True


def _extract_hybrid(pdf_path: str, source_folder: str = "") -> list:
    """
    逐页混合提取：每页先尝试 pdfplumber 文本，失败或无文本时再 OCR。
    解决其他电脑上整 PDF 被误判、OCR 缩图、分类过严等问题。
    """
    invoices = []
    skipped_pages = 0
    ocr = None
    ocr_usable = not _ocr_failed
    fname = os.path.basename(pdf_path)

    try:
        with pdfplumber.open(pdf_path) as pdf:
            doc = fitz.open(pdf_path)
            try:
                total_pages = len(pdf.pages)
                for page_num, page in enumerate(pdf.pages):
                    page_label = f"{page_num + 1}/{total_pages}"
                    _emit_progress(
                        f"  [OCR] {fname} 正在识别第 {page_label} 页...",
                        page_progress=int(page_num / total_pages * 100) if total_pages else 0,
                    )

                    full_text = page.extract_text() or ""
                    table_text = full_text

                    tables = page.extract_tables()
                    if tables:
                        cells = []
                        for table in tables:
                            for row in table:
                                cells.extend(cell for cell in row if cell)
                        if cells:
                            table_text = full_text + "\n" + "\n".join(cells)

                    if full_text.strip():
                        page_type = _classify_page(full_text, source='text')
                        if page_type == 'non_invoice':
                            skipped_pages += 1
                            continue
                        if page_type == 'invoice':
                            info = _build_invoice_from_text(
                                full_text, table_text, pdf_path, source_folder, "pdfplumber"
                            )
                            if info.invoice_number:
                                invoices.append(info)
                            continue

                    if not ocr_usable:
                        continue

                    if ocr is None:
                        try:
                            ocr = _get_ocr()
                        except RuntimeError:
                            ocr_usable = False
                            break

                    ocr_text = _ocr_page_text(ocr, doc[page_num], page_label)
                    if not ocr_text.strip():
                        continue

                    page_type = _classify_page(ocr_text, source='ocr')
                    if page_type != 'invoice':
                        skipped_pages += 1
                        continue

                    info = _build_invoice_from_text(
                        ocr_text, ocr_text, pdf_path, source_folder, "OCR"
                    )
                    if info.invoice_number:
                        invoices.append(info)
                        _emit_progress(
                            f"  [OCR] {fname} 第 {page_label} 页 ✓ "
                            f"发票 {info.invoice_number} ¥{info.total_with_tax or '?'}"
                        )
            finally:
                doc.close()

        if skipped_pages > 0:
            _emit_progress(
                f"  [过滤] '{fname}' 跳过 {skipped_pages}/{total_pages} 个非发票页面"
            )
        if invoices:
            _emit_progress(f"  [完成] '{fname}' 共识别 {len(invoices)} 张发票")
    except Exception as e:
        sys.stderr.write(f"  [错误] 解析 '{fname}' 时出错: {e}\n")
        sys.stderr.flush()

    if not invoices and ocr_usable is False and _ocr_failed:
        sys.stderr.write(
            f"  [跳过] '{fname}' 需要 OCR 识别，但 PaddleOCR 不可用\n"
        )
        sys.stderr.flush()

    return invoices


def extract_invoice_from_image(image_path: str, source_folder: str = "") -> list:
    """对单张发票图片（JPG/PNG 等）执行 OCR 识别。"""
    invoices = []
    fname = os.path.basename(image_path)

    if _ocr_failed:
        sys.stderr.write(f"  [跳过] '{fname}' 为图片文件，但 PaddleOCR 不可用\n")
        sys.stderr.flush()
        return invoices

    try:
        ocr = _get_ocr()
        _emit_progress(f"  [图片] '{fname}' → OCR 识别中...")
        ocr_text = _run_ocr_on_image(ocr, image_path)
        if not ocr_text.strip():
            return invoices

        page_type = _classify_page(ocr_text, source='ocr')
        if page_type != 'invoice':
            sys.stderr.write(f"  [过滤] '{fname}' 未识别为发票页面\n")
            sys.stderr.flush()
            return invoices

        info = _build_invoice_from_text(
            ocr_text, ocr_text, image_path, source_folder, "OCR"
        )
        if info.invoice_number:
            invoices.append(info)
            _emit_progress(
                f"  [图片] '{fname}' ✓ "
                f"发票 {info.invoice_number} ¥{info.total_with_tax or '?'}"
            )
        if invoices:
            _emit_progress(f"  [完成] '{fname}' 共识别 {len(invoices)} 张发票")
    except RuntimeError:
        pass
    except Exception as e:
        sys.stderr.write(f"  [错误] 图片识别 '{fname}' 时出错: {e}\n")
        sys.stderr.flush()

    return invoices


def extract_all_invoices(file_path: str, source_folder: str = "") -> list:
    """
    智能提取文件中的所有发票。
    支持 PDF（逐页混合策略）和图片文件（直接 OCR）。
    """
    if is_image_file(file_path):
        return extract_invoice_from_image(file_path, source_folder)

    fname = os.path.basename(file_path)

    if _is_image_based_pdf(file_path):
        if _ocr_failed:
            sys.stderr.write(f"  [跳过] '{fname}' 为图片型 PDF，但 PaddleOCR 不可用\n")
            sys.stderr.flush()
            return []
        print(f"  [图片型] '{fname}' → OCR 识别...")
        _emit_progress(f"  [图片型] '{fname}' → 开始逐页 OCR（共需识别每一页，请耐心等待）")
    else:
        print(f"  [混合] '{fname}' → 文本 + OCR 逐页识别...")
        _emit_progress(f"  [混合] '{fname}' → 开始逐页识别...")

    return _extract_hybrid(file_path, source_folder)


def _discover_folders(base_dir: str) -> dict:
    """
    递归扫描目录结构，发现所有发票文件（PDF / 图片）所在的文件夹。

    返回: { "相对路径或.": [文件路径列表], ... }
      - base_dir 下直接有文件 → {".": [文件列表]}
      - 子目录中的文件 → {"子文件夹": [...], "子文件夹/嵌套": [...], ...}
    """
    folders = {}
    base_dir = os.path.abspath(base_dir)

    if not os.path.isdir(base_dir):
        return folders

    for root, dirnames, filenames in os.walk(base_dir):
        dirnames[:] = sorted(
            d for d in dirnames
            if not d.startswith('.') and d != 'excel' and d != '.chunks'
        )
        invoice_files = sorted(
            os.path.join(root, f)
            for f in filenames
            if is_supported_invoice_file(f)
        )
        if not invoice_files:
            continue

        rel = os.path.relpath(root, base_dir)
        key = '.' if rel == '.' else rel.replace('\\', '/')
        folders.setdefault(key, []).extend(invoice_files)

    return folders


def _sanitize_sheet_name(name: str, used_names: set) -> str:
    """生成合法的 Excel 工作表名（最长 31 字符，不含非法字符）。"""
    if name == '.':
        name = '根目录'
    name = re.sub(r'[\\/*?:\[\]]', '_', name.replace('/', '_')).strip()
    if not name:
        name = '文件夹'
    base = name[:31]
    candidate = base
    suffix = 1
    while candidate in used_names:
        tail = f'_{suffix}'
        candidate = base[:31 - len(tail)] + tail
        suffix += 1
    used_names.add(candidate)
    return candidate


def _write_detail_sheet(ws, invoice_list, styles, show_folder_col=False):
    """写入发票明细表。"""
    header_font, header_fill, header_align, cell_font, cell_align, money_align, thin_border = styles

    headers = [
        "序号", "文件名", "发票号码", "开票日期",
        "购买方名称", "购买方（仅人名）", "销售方名称",
        "合计金额（不含税）", "价税合计金额",
        "手机号码", "计费时段",
    ]
    if show_folder_col:
        headers.insert(1, "来源文件夹")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, inv in enumerate(invoice_list, 2):
        buyer_clean = _buyer_name_base(inv.buyer_name)
        values = [
            row_idx - 1, inv.file_name,
            inv.invoice_number, inv.invoice_date,
            inv.buyer_name, buyer_clean, inv.seller_name,
            _safe_float(inv.subtotal),
            _safe_float(inv.total_with_tax),
            inv.phone_number, inv.billing_period,
        ]
        if show_folder_col:
            values.insert(1, inv.source_folder)

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            money_cols = (9, 10) if show_folder_col else (8, 9)
            if col_idx in money_cols:
                cell.alignment = money_align
                cell.number_format = '¥#,##0.00'
            else:
                cell.alignment = cell_align

    col_widths = ([6, 20] if show_folder_col else [6]) + [48, 26, 14, 32, 20, 40, 20, 20, 18, 14]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    if invoice_list:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(invoice_list) + 1}"


def _write_expense_name_sheet(ws, invoice_list, styles, show_folder_col=False):
    """写入报销文件名表。"""
    header_font, header_fill, header_align, cell_font, cell_align, money_align, thin_border = styles

    headers = [
        "序号", "报销文件", "发票号码", "开票日期",
        "购买方名称", "购买方（仅人名）", "销售方名称",
        "合计金额（不含税）", "价税合计金额",
        "手机号码", "计费时段",
    ]
    if show_folder_col:
        headers.insert(1, "来源文件夹")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, inv in enumerate(invoice_list, 2):
        billing_label = ""
        if inv.billing_period and len(inv.billing_period) == 7:
            parts = inv.billing_period.split('-')
            if len(parts) == 2:
                billing_label = f"{int(parts[0])}年{int(parts[1])}月"

        buyer_clean = _buyer_name_base(inv.buyer_name)
        expense_name = f"{buyer_clean}{billing_label}话费{inv.total_with_tax or '0'}"

        values = [
            row_idx - 1, expense_name,
            inv.invoice_number, inv.invoice_date,
            inv.buyer_name, buyer_clean, inv.seller_name,
            _safe_float(inv.subtotal),
            _safe_float(inv.total_with_tax),
            inv.phone_number, inv.billing_period,
        ]
        if show_folder_col:
            values.insert(1, inv.source_folder)

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
            money_cols = (9, 10) if show_folder_col else (8, 9)
            if col_idx in money_cols:
                cell.alignment = money_align
                cell.number_format = '¥#,##0.00'
            else:
                cell.alignment = cell_align

    col_widths = ([6, 20] if show_folder_col else [6]) + [52, 26, 14, 32, 40, 20, 20, 18, 14]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    if invoice_list:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(invoice_list) + 1}"


def write_to_excel(invoice_list: list, output_path: str, show_folder_col: bool = False,
                   folder_groups: dict = None):
    """
    将提取的发票信息写入 Excel 文件。

    folder_groups: { "文件夹名": [InvoiceInfo, ...], ... }
      多文件夹合并模式下，每个文件夹单独一个明细 Sheet；
      同时保留「报销汇总」「报销文件名」两个总表。
    """
    wb = openpyxl.Workbook()

    # ====== 样式 ======
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='微软雅黑', size=10)
    cell_align = Alignment(horizontal='center', vertical='center')
    money_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    styles = (header_font, header_fill, header_align, cell_font, cell_align, money_align, thin_border)

    use_folder_sheets = folder_groups and len(folder_groups) > 1

    if use_folder_sheets:
        # 多文件夹：每个文件夹一个明细 Sheet
        used_names = set()
        first = True
        for folder_name in sorted(folder_groups.keys()):
            sheet_title = _sanitize_sheet_name(folder_name, used_names)
            if first:
                ws = wb.active
                ws.title = sheet_title
                first = False
            else:
                ws = wb.create_sheet(sheet_title)
            _write_detail_sheet(ws, folder_groups[folder_name], styles, show_folder_col=False)
    else:
        # 单文件夹：原有「发票明细」Sheet
        ws = wb.active
        ws.title = "发票明细"
        _write_detail_sheet(ws, invoice_list, styles, show_folder_col=show_folder_col)

    # ====== 报销汇总 ======
    ws2 = wb.create_sheet("报销汇总")

    name_totals = {}
    name_counts = {}
    grand_total = 0
    for inv in invoice_list:
        name_base = _buyer_name_base(inv.buyer_name) or "未知"
        amount = _safe_float(inv.total_with_tax)
        name_totals[name_base] = name_totals.get(name_base, 0) + amount
        name_counts[name_base] = name_counts.get(name_base, 0) + 1
        grand_total += amount

    ws2.cell(row=1, column=1, value="发票数量：").font = Font(name='微软雅黑', size=11, bold=True)
    ws2.cell(row=1, column=2, value=f"{len(invoice_list)} 张").font = Font(name='微软雅黑', size=11)

    summary_headers = ["报销人", "金额", "发票张数"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 4
    for name, amount in name_totals.items():
        count = name_counts.get(name, 0)
        ws2.cell(row=row, column=1, value=name).font = cell_font
        ws2.cell(row=row, column=2, value=amount).font = cell_font
        ws2.cell(row=row, column=2).number_format = '¥#,##0.00'
        ws2.cell(row=row, column=3, value=count).font = cell_font
        for col in (1, 2, 3):
            ws2.cell(row=row, column=col).border = thin_border
        row += 1

    ws2.cell(row=row, column=1, value="总计").font = Font(name='微软雅黑', size=11, bold=True)
    ws2.cell(row=row, column=2, value=grand_total).font = Font(name='微软雅黑', size=11, bold=True)
    ws2.cell(row=row, column=2).number_format = '¥#,##0.00'
    for col in (1, 2, 3):
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).fill = PatternFill(
            start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 12

    # ====== 报销文件名 ======
    ws3 = wb.create_sheet("报销文件名")
    _write_expense_name_sheet(ws3, invoice_list, styles, show_folder_col=False)

    wb.save(output_path)
    print(f"\n  ✅ 已保存到: {output_path}")


def main():
    import datetime

    input_dir = CONFIG.get("input_dir") or SCRIPT_DIR
    excel_dir = os.path.join(SCRIPT_DIR, "excel")
    os.makedirs(excel_dir, exist_ok=True)

    print("=" * 60)
    print("  电子发票 PDF 自动识别工具 v2.0")
    print("  支持：文本型 PDF / 图片型 PDF（OCR）| 批量文件夹")
    print("=" * 60)
    print(f"\n  📂 扫描目录: {input_dir}")

    # ---- 发现所有包含 PDF 的文件夹 ----
    folders = _discover_folders(input_dir)

    if not folders:
        print("\n  ❌ 未找到任何发票文件（PDF / 图片）！")
        return

    # 统计
    total_file_count = sum(len(files) for files in folders.values())
    folder_names = list(folders.keys())

    if len(folders) == 1 and folder_names[0] == ".":
        print(f"  📁 单文件夹模式，共 {total_file_count} 个文件")
    else:
        print(f"  📁 发现 {len(folders)} 个文件夹，共 {total_file_count} 个文件：")
        for fname, files in folders.items():
            print(f"       · {fname} ({len(files)} 个文件)")

    print()

    # ---- 逐文件夹、逐文件解析 ----
    folder_invoices = {}  # { "文件夹名": [InvoiceInfo列表] }

    for folder_name, file_paths in folders.items():
        if folder_name != ".":
            print(f"  ▸ 文件夹: {folder_name}")
        else:
            print(f"  ▸ 根目录")

        invoices = []
        for idx, file_path in enumerate(file_paths, 1):
            fname = os.path.basename(file_path)
            print(f"     [{idx}/{len(file_paths)}] {fname}")
            results = extract_all_invoices(file_path, source_folder=folder_name)
            invoices.extend(results)

            if results:
                for inv in results:
                    print(f"          ✓ [{inv.extraction_method}] "
                          f"{inv.invoice_number}  ¥{inv.total_with_tax}  "
                          f"{inv.buyer_name} → {inv.seller_name}"
                          f"  📱{inv.phone_number or '?'}  📅{inv.billing_period or '?'}")
            else:
                print(f"          ✗ 未能提取到任何发票信息")
            print()

        folder_invoices[folder_name] = invoices
        folder_success = sum(1 for inv in invoices if inv.invoice_number)
        print(f"     📊 {folder_name}: {folder_success}/{len(invoices)} 张提取成功\n")

    # 总计
    all_invoices = [inv for invs in folder_invoices.values() for inv in invs]
    success = sum(1 for inv in all_invoices if inv.invoice_number)
    print(f"  📊 总识别结果: {success}/{len(all_invoices)} 张发票提取成功")

    # ---- 询问合并还是分开 ----
    is_multi_folder = len(folder_invoices) > 1 and list(folder_invoices.keys()) != ["."]

    if is_multi_folder:
        print(f"\n  🤔 检测到 {len(folder_invoices)} 个文件夹，如何输出？")
        print(f"     [1] 合并为一个 Excel 文件")
        print(f"     [2] 每个文件夹单独生成 Excel")
        merge_choice = input("     请选择 (1/2): ").strip()
        merge = (merge_choice == '1')
    else:
        merge = True  # 单文件夹默认合并（即正常输出）

    # ---- 确定输出文件路径 ----
    if merge:
        # 合并为一个 Excel
        default_output = os.path.join(excel_dir, "发票识别结果.xlsx")
        output_file = _resolve_output_path(default_output, "发票识别结果")
        write_to_excel(
            all_invoices, output_file,
            show_folder_col=False,
            folder_groups=folder_invoices if is_multi_folder else None,
        )
    else:
        # 每个文件夹单独输出
        for folder_name, invoices in folder_invoices.items():
            safe_name = folder_name if folder_name != "." else "根目录"
            default_output = os.path.join(excel_dir, f"发票识别结果_{safe_name}.xlsx")
            output_file = _resolve_output_path(default_output, f"发票识别结果_{safe_name}")
            write_to_excel(invoices, output_file)

    # ---- 汇总 ----
    print("\n" + "=" * 60)
    print("  提取汇总")
    print("=" * 60)
    total = 0
    for inv in all_invoices:
        amt = _safe_float(inv.total_with_tax)
        total += amt
        status = "✓" if inv.invoice_number else "✗"
        src = f"[{inv.source_folder}]" if inv.source_folder and inv.source_folder != "." else ""
        print(f"  {status} {inv.invoice_number:>26s}  {inv.invoice_date or '-':10s}  "
              f"¥{inv.total_with_tax or '0':>8s}  {inv.buyer_name or '-':12s}  {src}")
    print(f"  {'─' * 75}")
    print(f"  总计: {success} 张发票，¥{total:,.2f}")


def _resolve_output_path(default_path: str, base_name: str) -> str:
    """判断文件是否已存在，询问用户覆盖或新建。"""
    import datetime

    if os.path.exists(default_path):
        print(f"\n  ⚠ 已存在: {default_path}")
        print(f"     [1] 覆盖原文件")
        print(f"     [2] 新建文件（带时间戳）")
        choice = input("     请选择 (1/2): ").strip()

        if choice == '1':
            print(f"     → 将覆盖原文件")
            return default_path
        else:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            new_path = os.path.join(
                os.path.dirname(default_path),
                f"{base_name}_{timestamp}.xlsx"
            )
            print(f"     → 将新建文件: {os.path.basename(new_path)}")
            return new_path
    else:
        print(f"\n  🆕 首次识别，新建 Excel 文件")
        return default_path


if __name__ == '__main__':
    main()
